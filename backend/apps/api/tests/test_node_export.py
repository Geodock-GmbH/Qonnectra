"""Tests for node structure Excel export service and API endpoint."""

import io
import uuid

import openpyxl
import pytest
from apps.api.models import (
    AttributesComponentStructure,
    AttributesComponentType,
    NodeSlotConfiguration,
    NodeStructure,
)
from apps.api.services import generate_node_structure_excel
from apps.api.tests.factories import NodeFactory
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient

User = get_user_model()


@pytest.mark.django_db
class TestGenerateNodeStructureExcel:
    """Tests for the node structure Excel export service."""

    def setup_method(self):
        """Create a node with slot configuration for export tests."""
        self.node = NodeFactory(name="Test Node")
        self.slot_config = NodeSlotConfiguration.objects.create(
            uuid_node=self.node,
            side="A",
            total_slots=12,
        )

    def test_returns_valid_excel_response(self):
        """Export returns an HttpResponse with Excel content type."""
        response = generate_node_structure_excel(self.node.uuid)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_has_one_sheet_per_side(self):
        """Each SlotConfiguration becomes its own Excel sheet."""
        NodeSlotConfiguration.objects.create(
            uuid_node=self.node,
            side="B",
            total_slots=6,
        )

        response = generate_node_structure_excel(self.node.uuid)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        assert len(wb.sheetnames) == 2
        assert "A" in wb.sheetnames
        assert "B" in wb.sheetnames

    def test_node_not_found_returns_none(self):
        """Non-existent node UUID returns None (view handles 404)."""
        result = generate_node_structure_excel(uuid.uuid4())
        assert result is None

    def test_node_with_no_configs_returns_empty_sheet(self):
        """Node with no slot configurations returns workbook with 'No Data' sheet."""
        empty_node = NodeFactory(name="Empty Node")
        response = generate_node_structure_excel(empty_node.uuid)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))

        assert wb.sheetnames == ["No Data"]

    def test_component_generates_port_rows(self):
        """A component with ports generates one row per port."""
        comp_type = AttributesComponentType.objects.create(
            component_type="Spleisskassette SC",
            occupied_slots=1,
        )
        for i in range(1, 4):
            AttributesComponentStructure.objects.create(
                component_type=comp_type, in_or_out="in", port=i
            )
            AttributesComponentStructure.objects.create(
                component_type=comp_type, in_or_out="out", port=i
            )

        NodeStructure.objects.create(
            uuid_node=self.node,
            slot_configuration=self.slot_config,
            component_type=comp_type,
            slot_start=1,
            slot_end=1,
            purpose=NodeStructure.Purpose.COMPONENT,
        )

        response = generate_node_structure_excel(self.node.uuid)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["A"]

        # Row 4 = header, rows 5-7 = 3 port rows
        assert ws.cell(row=5, column=5).value == 1  # Port 1
        assert ws.cell(row=6, column=5).value == 2  # Port 2
        assert ws.cell(row=7, column=5).value == 3  # Port 3
        assert ws.cell(row=5, column=3).value == "Spleisskassette SC"

    def test_empty_slot_shows_purpose(self):
        """Empty/reserve slots show purpose label."""
        NodeStructure.objects.create(
            uuid_node=self.node,
            slot_configuration=self.slot_config,
            slot_start=1,
            slot_end=1,
            purpose=NodeStructure.Purpose.RESERVE,
        )

        response = generate_node_structure_excel(self.node.uuid)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb["A"]

        assert ws.cell(row=5, column=3).value == "Reserve"


@pytest.mark.django_db
class TestNodeStructureExportView:
    """Tests for the export API endpoint."""

    def setup_method(self):
        """Create an authenticated client and test node for export endpoint tests."""
        self.client = APIClient()
        self.user = User.objects.create_user(
            username="testuser", password="testpass123"
        )
        self.client.force_authenticate(user=self.user)
        self.node = NodeFactory(name="Test Node")

    def test_authenticated_user_can_download(self):
        """Test that an authenticated user can download an Excel export."""
        NodeSlotConfiguration.objects.create(
            uuid_node=self.node, side="A", total_slots=12
        )
        response = self.client.get(f"/api/v1/node-export/excel/{self.node.uuid}/")
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_unauthenticated_returns_401(self):
        """Test that unauthenticated requests return 401."""
        self.client.force_authenticate(user=None)
        response = self.client.get(f"/api/v1/node-export/excel/{self.node.uuid}/")
        assert response.status_code == 401

    def test_nonexistent_node_returns_404(self):
        """Test that a non-existent node returns 404."""
        response = self.client.get(f"/api/v1/node-export/excel/{uuid.uuid4()}/")
        assert response.status_code == 404
