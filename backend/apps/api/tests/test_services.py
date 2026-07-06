"""
Unit tests for service functions in apps/api/services.py.

Tests cover:
- import_conduits_from_excel: Excel import validation and processing
- generate_conduit_import_template: Template generation
- rename_feature_folder: Feature folder renaming
- move_file_to_feature: File moving between features
- handle_qgis_file: QGIS file handling
- convert_qgs_to_postgres: QGS datasource conversion
"""

import io
import zipfile
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from apps.api.models import Conduit, FeatureFiles, Microduct, Node
from apps.api.services import (
    _build_postgres_datasource,
    _extract_filename_from_ogr_source,
    _extract_layer_name_from_gpkg_source,
    _postgres_type_to_pandas,
    _rewrite_file_datasources,
    build_inquiry_export_zip,
    convert_qgs_to_postgres,
    generate_conduit_import_template,
    handle_qgis_file,
    import_conduits_from_excel,
    move_file_to_feature,
    rename_feature_folder,
    repackage_qgz,
)
from django.utils.translation import activate

from .factories import (
    AddressFactory,
    AreaFactory,
    CableTypeFactory,
    CompanyFactory,
    ConduitFactory,
    ConduitTypeColorMappingFactory,
    ConduitTypeFactory,
    FlagFactory,
    MicroductColorFactory,
    MicroductStatusFactory,
    NetworkLevelFactory,
    NodeFactory,
    NodeTypeFactory,
    PhaseFactory,
    PipelineInquiryAreaFactory,
    PipelineRecordFactory,
    ProjectFactory,
    StatusDevelopmentFactory,
    StatusFactory,
    StoragePreferencesFactory,
    SurfaceFactory,
    TrenchConduitConnectionFactory,
    TrenchFactory,
)


@pytest.fixture
def excel_file_with_valid_data(db):
    """Create an Excel file with valid conduit data."""
    project = ProjectFactory(project="TestProject")
    flag = FlagFactory(flag="TestFlag")
    conduit_type = ConduitTypeFactory(conduit_type="12x10/6")
    status = StatusFactory(status="geplant")
    network_level = NetworkLevelFactory(network_level="Hausanschluss-Ebene")
    owner = CompanyFactory(company="Geodock")

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = [
        "Name",
        "Type",
        "Outer Conduit",
        "Status",
        "Network Level",
        "Owner",
        "Constructor",
        "Manufacturer",
        "Date",
        "Project",
        "Flag",
    ]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col, value=header)

    data_row = [
        "RV1.1.1",
        "12x10/6",
        "",
        "geplant",
        "Hausanschluss-Ebene",
        "Geodock",
        "",
        "",
        "2025-01-01",
        "TestProject",
        "TestFlag",
    ]
    for col, value in enumerate(data_row, start=1):
        sheet.cell(row=2, column=col, value=value)

    file_buffer = io.BytesIO()
    workbook.save(file_buffer)
    file_buffer.seek(0)

    return {
        "file": file_buffer,
        "project": project,
        "flag": flag,
        "conduit_type": conduit_type,
        "status": status,
        "network_level": network_level,
        "owner": owner,
    }


@pytest.fixture
def conduit_type_with_colors(db):
    """Create a conduit type with color mappings for microduct creation."""
    colors = []
    color_names = [
        ("rot", "red", "#dc2626"),
        ("grün", "green", "#16a34a"),
        ("blau", "blue", "#2563eb"),
    ]
    for name_de, name_en, hex_code in color_names:
        colors.append(
            MicroductColorFactory(
                name_de=name_de,
                name_en=name_en,
                hex_code=hex_code,
            )
        )

    conduit_type = ConduitTypeFactory(conduit_type="3x10/6", conduit_count=3)
    for i, color in enumerate(colors, 1):
        ConduitTypeColorMappingFactory(
            conduit_type=conduit_type,
            position=i,
            color=color,
        )
    return conduit_type


@pytest.mark.django_db
class TestImportConduitsFromExcel:
    """Tests for the import_conduits_from_excel service function."""

    def test_import_valid_excel_file(self, excel_file_with_valid_data):
        """Test successful import of valid Excel file."""
        activate("en")
        result = import_conduits_from_excel(excel_file_with_valid_data["file"])

        assert result["success"] is True
        assert result["created_count"] == 1

        conduit = Conduit.objects.get(name="RV1.1.1")
        assert conduit.project == excel_file_with_valid_data["project"]
        assert conduit.flag == excel_file_with_valid_data["flag"]

    def test_import_creates_microducts_for_conduit_type_with_colors(
        self, db, conduit_type_with_colors
    ):
        """Test that microducts are created when conduit type has color mappings."""
        activate("en")
        ProjectFactory(project="TestProject2")
        FlagFactory(flag="TestFlag2")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = [
            "Name",
            "Type",
            "Outer Conduit",
            "Status",
            "Network Level",
            "Owner",
            "Constructor",
            "Manufacturer",
            "Date",
            "Project",
            "Flag",
        ]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        data_row = [
            "TestConduit",
            "3x10/6",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "TestProject2",
            "TestFlag2",
        ]
        for col, value in enumerate(data_row, start=1):
            sheet.cell(row=2, column=col, value=value)

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is True
        conduit = Conduit.objects.get(name="TestConduit")
        microducts = Microduct.objects.filter(uuid_conduit=conduit)
        assert microducts.count() == 3

    def test_import_file_too_large(self, db):
        """Test rejection of file exceeding size limit."""
        activate("en")
        large_file = io.BytesIO(b"x" * (10 * 1024 * 1024 + 1))
        result = import_conduits_from_excel(large_file, max_file_size=10 * 1024 * 1024)

        assert result["success"] is False
        assert any("too large" in err.lower() for err in result["errors"])

    def test_import_invalid_excel_file(self, db):
        """Test rejection of invalid Excel file."""
        activate("en")
        invalid_file = io.BytesIO(b"not an excel file")
        result = import_conduits_from_excel(invalid_file)

        assert result["success"] is False
        assert any("invalid excel" in err.lower() for err in result["errors"])

    def test_import_missing_name_column(self, db):
        """Test rejection when required Name column is missing."""
        activate("en")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="Type")
        sheet.cell(row=1, column=2, value="Project")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("name" in err.lower() for err in result["errors"])

    def test_import_empty_headers(self, db):
        """Test rejection when no headers found."""
        activate("en")
        workbook = openpyxl.Workbook()

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("header" in err.lower() for err in result["errors"])

    def test_import_duplicate_conduit_name(self, db):
        """Test rejection when conduit name already exists."""
        activate("en")
        project = ProjectFactory(project="TestProject3")
        flag = FlagFactory(flag="TestFlag3")
        conduit_type = ConduitTypeFactory()
        Conduit.objects.create(
            name="ExistingConduit",
            project=project,
            flag=flag,
            conduit_type=conduit_type,
        )

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Project", "Flag"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=2, column=1, value="ExistingConduit")
        sheet.cell(row=2, column=2, value="TestProject3")
        sheet.cell(row=2, column=3, value="TestFlag3")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("already exists" in err.lower() for err in result["errors"])

    def test_import_missing_row_name(self, db):
        """Test error when row is missing name."""
        activate("en")
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Type"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=2, column=1, value="")
        sheet.cell(row=2, column=2, value="SomeType")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("required" in str(err).lower() for err in result["errors"])

    def test_import_nonexistent_project(self, db):
        """Test error when project doesn't exist."""
        activate("en")
        FlagFactory(flag="TestFlag4")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Project", "Flag"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=2, column=1, value="TestConduit")
        sheet.cell(row=2, column=2, value="NonExistentProject")
        sheet.cell(row=2, column=3, value="TestFlag4")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("not found" in str(err).lower() for err in result["errors"])

    def test_import_nonexistent_owner_company(self, db):
        """Test error when owner company doesn't exist."""
        activate("en")
        ProjectFactory(project="TestProject5")
        FlagFactory(flag="TestFlag5")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Owner", "Project", "Flag"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=2, column=1, value="TestConduit")
        sheet.cell(row=2, column=2, value="NonExistentCompany")
        sheet.cell(row=2, column=3, value="TestProject5")
        sheet.cell(row=2, column=4, value="TestFlag5")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is False
        assert any("not found" in str(err).lower() for err in result["errors"])

    def test_import_warns_about_unrecognized_columns(self, db):
        """Test that unrecognized columns generate warnings."""
        activate("en")
        ProjectFactory(project="TestProject6")
        FlagFactory(flag="TestFlag6")
        ConduitTypeFactory(conduit_type="TestType")

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        headers = ["Name", "Type", "UnknownColumn", "Project", "Flag"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        sheet.cell(row=2, column=1, value="TestConduit")
        sheet.cell(row=2, column=2, value="TestType")
        sheet.cell(row=2, column=3, value="SomeValue")
        sheet.cell(row=2, column=4, value="TestProject6")
        sheet.cell(row=2, column=5, value="TestFlag6")

        file_buffer = io.BytesIO()
        workbook.save(file_buffer)
        file_buffer.seek(0)

        result = import_conduits_from_excel(file_buffer)

        assert result["success"] is True
        assert "warnings" in result
        assert any("ignored" in str(w).lower() for w in result["warnings"])


@pytest.mark.django_db
class TestGenerateConduitImportTemplate:
    """Tests for the generate_conduit_import_template service function."""

    def test_generates_valid_excel_response(self):
        """Test that template generates a valid Excel HTTP response."""
        activate("en")
        response = generate_conduit_import_template()

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "conduit_import_template.xlsx" in response["Content-Disposition"]

    def test_template_has_correct_headers(self):
        """Test that template has all required headers."""
        activate("en")
        response = generate_conduit_import_template()

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        expected_headers = [
            "Name",
            "Type",
            "Outer Conduit",
            "Status",
            "Network Level",
            "Owner",
            "Constructor",
            "Manufacturer",
            "Date",
            "Project",
            "Flag",
        ]
        actual_headers = [sheet.cell(row=1, column=i).value for i in range(1, 12)]

        assert actual_headers == expected_headers

    def test_template_has_example_row(self):
        """Test that template has an example data row."""
        activate("en")
        response = generate_conduit_import_template()

        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active

        name_value = sheet.cell(row=2, column=1).value
        assert name_value == "RV1.1.1"


class TestHandleQgisFile:
    """Tests for the handle_qgis_file service function."""

    def test_handle_qgs_file(self):
        """Test handling of plain QGS file."""
        qgs_content = b'<?xml version="1.0"?><qgis></qgis>'
        result_content, is_qgz = handle_qgis_file(qgs_content, "test.qgs")

        assert result_content == qgs_content
        assert is_qgz is False

    def test_handle_qgz_file(self):
        """Test handling of QGZ archive file."""
        qgs_content = b'<?xml version="1.0"?><qgis></qgis>'

        qgz_buffer = io.BytesIO()
        with zipfile.ZipFile(qgz_buffer, "w") as zf:
            zf.writestr("project.qgs", qgs_content)
        qgz_buffer.seek(0)

        result_content, is_qgz = handle_qgis_file(qgz_buffer.getvalue(), "test.qgz")

        assert result_content == qgs_content
        assert is_qgz is True

    def test_handle_qgz_without_qgs_raises_error(self):
        """Test that QGZ without QGS file raises ValueError."""
        qgz_buffer = io.BytesIO()
        with zipfile.ZipFile(qgz_buffer, "w") as zf:
            zf.writestr("other.txt", b"some content")
        qgz_buffer.seek(0)

        with pytest.raises(ValueError, match="No .qgs file found"):
            handle_qgis_file(qgz_buffer.getvalue(), "test.qgz")


class TestRepackageQgz:
    """Tests for the repackage_qgz service function."""

    def test_repackage_qgz_preserves_other_files(self):
        """Test that repackage preserves non-QGS files."""
        original_qgs = b'<?xml version="1.0"?><qgis>original</qgis>'
        modified_qgs = b'<?xml version="1.0"?><qgis>modified</qgis>'
        other_content = b"other file content"

        qgz_buffer = io.BytesIO()
        with zipfile.ZipFile(qgz_buffer, "w") as zf:
            zf.writestr("project.qgs", original_qgs)
            zf.writestr("data.txt", other_content)
        qgz_buffer.seek(0)

        result = repackage_qgz(modified_qgs, qgz_buffer.getvalue())

        with zipfile.ZipFile(io.BytesIO(result), "r") as zf:
            assert zf.read("project.qgs") == modified_qgs
            assert zf.read("data.txt") == other_content


class TestConvertQgsToPostgres:
    """Tests for the convert_qgs_to_postgres service function."""

    def test_converts_gpkg_layers_to_postgres(self):
        """Test conversion of GeoPackage datasources to PostgreSQL."""
        qgs_content = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="ogr" source="./schema.gpkg|layername=trench"/>
            <maplayer>
                <provider>ogr</provider>
                <datasource>./schema.gpkg|layername=node</datasource>
            </maplayer>
        </qgis>
        """

        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.QGIS_PG_SERVICE_NAME = "qonnectra"
            mock_settings.DEFAULT_SRID = 25832

            result = convert_qgs_to_postgres(qgs_content)

        assert b"postgres" in result
        assert b"service='qonnectra'" in result

    def test_preserves_non_gpkg_layers(self):
        """Test that non-GeoPackage layers are preserved."""
        qgs_content = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="wms" source="https://wms.example.com"/>
        </qgis>
        """

        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.QGIS_PG_SERVICE_NAME = "qonnectra"
            mock_settings.DEFAULT_SRID = 25832

            result = convert_qgs_to_postgres(qgs_content)

        assert b'providerKey="wms"' in result

    def test_rewrites_local_file_datasources(self):
        """Test that local file paths are rewritten to container paths."""
        qgs_content = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="ogr" source="/Users/john/data/survey.dxf"/>
            <maplayer>
                <provider>ogr</provider>
                <datasource>/Users/john/data/survey.dxf</datasource>
            </maplayer>
        </qgis>
        """

        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.QGIS_PG_SERVICE_NAME = "qonnectra"
            mock_settings.DEFAULT_SRID = 25832

            result = convert_qgs_to_postgres(
                qgs_content,
                data_filenames=["survey.dxf"],
                project_name="my-project",
            )

        assert b"/data/my-project/survey.dxf" in result
        assert b'providerKey="ogr"' in result  # stays ogr, not converted to postgres

    def test_mixed_gpkg_and_local_files(self):
        """Test that GeoPackage layers convert to PostgreSQL while local files get path-rewritten."""
        qgs_content = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="ogr" source="./schema.gpkg|layername=trench"/>
            <layer-tree-layer providerKey="ogr" source="/Users/john/data/survey.dxf"/>
            <maplayer>
                <provider>ogr</provider>
                <datasource>./schema.gpkg|layername=trench</datasource>
            </maplayer>
            <maplayer>
                <provider>ogr</provider>
                <datasource>/Users/john/data/survey.dxf</datasource>
            </maplayer>
        </qgis>
        """

        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.QGIS_PG_SERVICE_NAME = "qonnectra"
            mock_settings.DEFAULT_SRID = 25832

            result = convert_qgs_to_postgres(
                qgs_content,
                data_filenames=["survey.dxf"],
                project_name="infra",
            )

        # GeoPackage layer converted to PostgreSQL
        assert b"service='qonnectra'" in result
        assert b"postgres" in result
        # Local file layer path-rewritten
        assert b"/data/infra/survey.dxf" in result

    def test_no_rewriting_without_filenames(self):
        """Test backward compatibility — no file rewriting when data_filenames is None."""
        qgs_content = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="ogr" source="/Users/john/data/survey.dxf"/>
        </qgis>
        """

        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.QGIS_PG_SERVICE_NAME = "qonnectra"
            mock_settings.DEFAULT_SRID = 25832

            result = convert_qgs_to_postgres(qgs_content)

        # Without data_filenames, local file paths are left untouched
        assert b"/Users/john/data/survey.dxf" in result


class TestExtractLayerNameFromGpkgSource:
    """Tests for the _extract_layer_name_from_gpkg_source helper."""

    def test_extracts_layer_name(self):
        """Test extraction of layer name from source string."""
        source = "./schema.gpkg|layername=address"
        result = _extract_layer_name_from_gpkg_source(source)
        assert result == "address"

    def test_returns_none_for_non_gpkg_source(self):
        """Test returns None for non-GeoPackage source."""
        source = "https://wms.example.com"
        result = _extract_layer_name_from_gpkg_source(source)
        assert result is None


class TestBuildPostgresDatasource:
    """Tests for the _build_postgres_datasource helper."""

    def test_builds_geometry_layer_datasource(self):
        """Test datasource string for geometry layer."""
        with patch("apps.api.services.settings") as mock_settings:
            mock_settings.DEFAULT_SRID = 25832

            result = _build_postgres_datasource("trench", "qonnectra", 25832)

        assert "service='qonnectra'" in result
        assert "trench" in result
        assert "LineString" in result
        assert "geom" in result

    def test_builds_non_geometry_layer_datasource(self):
        """Test datasource string for non-geometry layer."""
        result = _build_postgres_datasource("conduit", "qonnectra", 25832)

        assert "service='qonnectra'" in result
        assert "conduit" in result
        assert "geometry" not in result.lower() or "type=" not in result

    def test_returns_none_for_unknown_layer(self):
        """Test returns None for unknown layer name."""
        result = _build_postgres_datasource("unknown_layer", "qonnectra", 25832)
        assert result is None


class TestPostgresTypeToPandas:
    """Tests for the _postgres_type_to_pandas helper."""

    def test_maps_common_types(self):
        """Test mapping of common PostgreSQL types."""
        assert _postgres_type_to_pandas("uuid", "uuid") == "object"
        assert _postgres_type_to_pandas("character varying", "varchar") == "object"
        assert _postgres_type_to_pandas("integer", "int4") == "Int64"
        assert _postgres_type_to_pandas("double precision", "float8") == "float64"
        assert _postgres_type_to_pandas("boolean", "bool") == "boolean"

    def test_returns_object_for_unknown_type(self):
        """Test returns object for unknown types."""
        assert _postgres_type_to_pandas("unknown_type", "unknown") == "object"


@pytest.mark.django_db
class TestRenameFeatureFolder:
    """Tests for the rename_feature_folder service function."""

    @patch("apps.api.services.LocalMediaStorage")
    def test_rename_folder_success(self, mock_storage_class):
        """Test successful folder rename when folder exists."""
        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})
        node = NodeFactory(name="OldNodeName", project=project)

        mock_storage = MagicMock()
        mock_storage.rename_folder.return_value = True
        mock_storage_class.return_value = mock_storage

        rename_feature_folder(node, "OldNodeName", "NewNodeName")

        mock_storage.rename_folder.assert_called_once()
        call_args = mock_storage.rename_folder.call_args[0]
        assert "OldNodeName" in call_args[0]
        assert "NewNodeName" in call_args[1]

    @patch("apps.api.services.LocalMediaStorage")
    def test_rename_folder_not_found(self, mock_storage_class):
        """Test rename when source folder doesn't exist (no files uploaded yet)."""
        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})
        node = NodeFactory(name="NodeWithoutFiles", project=project)

        mock_storage = MagicMock()
        mock_storage.rename_folder.return_value = False
        mock_storage_class.return_value = mock_storage

        rename_feature_folder(node, "NodeWithoutFiles", "NewNodeName")

        mock_storage.rename_folder.assert_called_once()

    @patch("apps.api.services.LocalMediaStorage")
    def test_rename_folder_updates_file_paths(self, mock_storage_class):
        """Test that file paths are updated in database when folder is renamed."""
        from apps.api.models import FeatureFiles, Node
        from django.contrib.contenttypes.models import ContentType

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})
        node = NodeFactory(name="OldNodeName", project=project)
        content_type = ContentType.objects.get_for_model(Node)

        file_record = FeatureFiles.objects.create(
            content_type=content_type,
            object_id=node.pk,
            file_name="test_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/OldNodeName/test_file.pdf"
        file_record.save()

        mock_storage = MagicMock()
        mock_storage.rename_folder.return_value = True
        mock_storage_class.return_value = mock_storage

        rename_feature_folder(node, "OldNodeName", "NewNodeName")

        file_record.refresh_from_db()
        assert "NewNodeName" in file_record.file_path.name
        assert "OldNodeName" not in file_record.file_path.name

    @patch("apps.api.services.LocalMediaStorage")
    def test_rename_folder_without_preferences(self, mock_storage_class):
        """Test folder rename when no storage preferences exist."""
        project = ProjectFactory(project="TestProject")
        node = NodeFactory(name="OldNodeName", project=project)

        mock_storage = MagicMock()
        mock_storage.rename_folder.return_value = True
        mock_storage_class.return_value = mock_storage

        rename_feature_folder(node, "OldNodeName", "NewNodeName")

        mock_storage.rename_folder.assert_called_once()
        call_args = mock_storage.rename_folder.call_args[0]
        assert "nodes" in call_args[0]

    @patch("apps.api.services.LocalMediaStorage")
    def test_rename_folder_for_trench(self, mock_storage_class):
        """Test folder rename for Trench model uses id_trench path."""
        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"trench": {"default": "trenches"}})
        trench = TrenchFactory(id_trench="TR-AAAAAAA", project=project)

        mock_storage = MagicMock()
        mock_storage.rename_folder.return_value = True
        mock_storage_class.return_value = mock_storage

        rename_feature_folder(trench, "TR-AAAAAAA", "TR-BBBBBBB")

        mock_storage.rename_folder.assert_called_once()
        call_args = mock_storage.rename_folder.call_args[0]
        assert "TR-AAAAAAA" in call_args[0]
        assert "TR-BBBBBBB" in call_args[1]


@pytest.mark.django_db
class TestMoveFileToFeature:
    """Tests for the move_file_to_feature service function."""

    @patch("apps.api.services.LocalMediaStorage")
    def test_move_file_success(self, mock_storage_class):
        """Test successful file move to another feature."""
        from django.contrib.contenttypes.models import ContentType

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})

        source_node = NodeFactory(name="SourceNode", project=project)
        target_node = NodeFactory(name="TargetNode", project=project)
        content_type = ContentType.objects.get_for_model(Node)

        file_record = FeatureFiles.objects.create(
            content_type=content_type,
            object_id=source_node.pk,
            file_name="test_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/SourceNode/test_file.pdf"
        file_record.save()

        mock_file = MagicMock()
        mock_file.read.return_value = b"file content"

        mock_storage = MagicMock()
        mock_storage.exists.side_effect = (
            lambda path: path == "TestProject/nodes/SourceNode/test_file.pdf"
        )
        mock_storage.open.return_value = mock_file
        mock_storage.save.return_value = "TestProject/nodes/TargetNode/test_file.pdf"
        mock_storage.delete.return_value = None
        mock_storage_class.return_value = mock_storage

        success, new_path, error = move_file_to_feature(
            file_record, target_node, content_type
        )

        assert success is True
        assert new_path is not None
        assert "TargetNode" in new_path
        assert error is None

    @patch("apps.api.services.LocalMediaStorage")
    def test_move_file_source_not_found(self, mock_storage_class):
        """Test move when source file doesn't exist on disk."""
        from django.contrib.contenttypes.models import ContentType

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})

        source_node = NodeFactory(name="SourceNode", project=project)
        target_node = NodeFactory(name="TargetNode", project=project)
        content_type = ContentType.objects.get_for_model(Node)

        file_record = FeatureFiles.objects.create(
            content_type=content_type,
            object_id=source_node.pk,
            file_name="nonexistent_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/SourceNode/nonexistent_file.pdf"
        file_record.save()

        mock_storage = MagicMock()
        mock_storage.exists.return_value = False
        mock_storage_class.return_value = mock_storage

        success, new_path, error = move_file_to_feature(
            file_record, target_node, content_type
        )

        assert success is True
        assert new_path is not None
        mock_storage.save.assert_not_called()
        mock_storage.delete.assert_not_called()

    @patch("apps.api.services.LocalMediaStorage")
    def test_move_file_target_exists(self, mock_storage_class):
        """Test move fails when file already exists at target path."""
        from django.contrib.contenttypes.models import ContentType

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})

        source_node = NodeFactory(name="SourceNode", project=project)
        target_node = NodeFactory(name="TargetNode", project=project)
        content_type = ContentType.objects.get_for_model(Node)

        file_record = FeatureFiles.objects.create(
            content_type=content_type,
            object_id=source_node.pk,
            file_name="test_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/SourceNode/test_file.pdf"
        file_record.save()

        mock_storage = MagicMock()
        mock_storage.exists.return_value = True
        mock_storage_class.return_value = mock_storage

        success, new_path, error = move_file_to_feature(
            file_record, target_node, content_type
        )

        assert success is False
        assert new_path is None
        assert "already exists" in error

    @patch("apps.api.services.LocalMediaStorage")
    def test_move_file_storage_error(self, mock_storage_class):
        """Test move handles storage errors gracefully."""
        from django.contrib.contenttypes.models import ContentType

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"node": {"default": "nodes"}})

        source_node = NodeFactory(name="SourceNode", project=project)
        target_node = NodeFactory(name="TargetNode", project=project)
        content_type = ContentType.objects.get_for_model(Node)

        file_record = FeatureFiles.objects.create(
            content_type=content_type,
            object_id=source_node.pk,
            file_name="test_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/SourceNode/test_file.pdf"
        file_record.save()

        mock_storage = MagicMock()
        mock_storage.exists.side_effect = [False, True]
        mock_storage.open.side_effect = Exception("Storage error")
        mock_storage_class.return_value = mock_storage

        success, new_path, error = move_file_to_feature(
            file_record, target_node, content_type
        )

        assert success is False
        assert new_path is None
        assert "Storage error" in error

    @patch("apps.api.services.LocalMediaStorage")
    def test_move_file_to_trench(self, mock_storage_class):
        """Test move file to a trench feature uses id_trench."""
        from django.contrib.contenttypes.models import ContentType
        from apps.api.models import Trench

        project = ProjectFactory(project="TestProject")
        StoragePreferencesFactory(folder_structure={"trench": {"default": "trenches"}})

        source_node = NodeFactory(name="SourceNode", project=project)
        target_trench = TrenchFactory(id_trench="TR-AAAAAAA", project=project)
        node_content_type = ContentType.objects.get_for_model(Node)
        trench_content_type = ContentType.objects.get_for_model(Trench)

        file_record = FeatureFiles.objects.create(
            content_type=node_content_type,
            object_id=source_node.pk,
            file_name="test_file",
            file_type="pdf",
        )
        file_record.file_path.name = "TestProject/nodes/SourceNode/test_file.pdf"
        file_record.save()

        mock_file = MagicMock()
        mock_file.read.return_value = b"file content"

        mock_storage = MagicMock()
        mock_storage.exists.side_effect = (
            lambda path: path == "TestProject/nodes/SourceNode/test_file.pdf"
        )
        mock_storage.open.return_value = mock_file
        mock_storage.save.return_value = "TestProject/trenches/TR-001/test_file.pdf"
        mock_storage.delete.return_value = None
        mock_storage_class.return_value = mock_storage

        success, new_path, error = move_file_to_feature(
            file_record, target_trench, trench_content_type
        )

        assert success is True
        assert new_path is not None
        assert "TR-AAAAAAA" in new_path
        assert error is None


class TestExtractFilenameFromOgrSource:
    """Tests for cross-platform filename extraction from OGR datasource strings."""

    def test_unix_path(self):
        """Extract filename from a standard Unix absolute path."""
        source = "/Users/john/data/survey.dxf"
        assert _extract_filename_from_ogr_source(source) == "survey.dxf"

    def test_unix_path_with_layername(self):
        """Extract filename when OGR |layername= parameter is present."""
        source = "/Users/john/data/buildings.gpkg|layername=floor_plans"
        assert _extract_filename_from_ogr_source(source) == "buildings.gpkg"

    def test_unix_path_with_layername_and_subset(self):
        """Extract filename when both |layername= and |subset= are present."""
        source = "/Users/maltethen/Downloads/170828-Bestand_BA 1.gpkg|layername=polylines|subset=layer IN ('BZV_Blattschnitt_Text') AND space=0 AND block=-1"
        assert _extract_filename_from_ogr_source(source) == "170828-Bestand_BA 1.gpkg"

    def test_windows_path(self):
        """Extract filename from a Windows backslash-separated path."""
        source = r"C:\Projects\old_data.shp"
        assert _extract_filename_from_ogr_source(source) == "old_data.shp"

    def test_windows_path_with_spaces(self):
        """Extract filename from a Windows path containing spaces."""
        source = r"D:\GIS Data\survey files\plan.dxf"
        assert _extract_filename_from_ogr_source(source) == "plan.dxf"

    def test_relative_path(self):
        """Extract filename from a relative Unix path."""
        source = "./data/survey.dxf"
        assert _extract_filename_from_ogr_source(source) == "survey.dxf"

    def test_filename_only(self):
        """Return the filename as-is when no directory prefix exists."""
        source = "survey.dxf"
        assert _extract_filename_from_ogr_source(source) == "survey.dxf"


class TestRewriteFileDatasources:
    """Tests for rewriting local file paths in QGS XML to container paths."""

    def test_rewrites_local_file_path_in_layer_tree(self):
        """Test rewriting a layer-tree-layer source attribute."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="ogr" source="/Users/john/data/survey.dxf"/>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        rewritten = _rewrite_file_datasources(root, ["survey.dxf"], "my-project")

        layer = root.find(".//layer-tree-layer")
        assert layer.get("source") == "/data/my-project/survey.dxf"
        assert layer.get("providerKey") == "ogr"
        assert rewritten == ["survey.dxf"]

    def test_rewrites_local_file_path_in_maplayer(self):
        """Test rewriting a maplayer datasource element."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <maplayer>
                <provider>ogr</provider>
                <datasource>/Users/john/data/survey.dxf</datasource>
            </maplayer>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        _rewrite_file_datasources(root, ["survey.dxf"], "my-project")

        ds = root.find(".//datasource")
        assert ds.text == "/data/my-project/survey.dxf"

    def test_preserves_ogr_parameters(self):
        """Test that |layername= and |subset= are preserved after rewriting."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <maplayer>
                <provider>ogr</provider>
                <datasource>/Users/maltethen/Downloads/file.gpkg|layername=polylines|subset=layer IN ('X')</datasource>
            </maplayer>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        _rewrite_file_datasources(root, ["file.gpkg"], "my-project")

        ds = root.find(".//datasource")
        assert (
            ds.text
            == "/data/my-project/file.gpkg|layername=polylines|subset=layer IN ('X')"
        )

    def test_skips_non_matching_filenames(self):
        """Test that layers with non-uploaded files are left unchanged."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <maplayer>
                <provider>ogr</provider>
                <datasource>/Users/john/data/not_uploaded.dxf</datasource>
            </maplayer>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        rewritten = _rewrite_file_datasources(root, ["other.dxf"], "my-project")

        ds = root.find(".//datasource")
        assert ds.text == "/Users/john/data/not_uploaded.dxf"
        assert rewritten == []

    def test_skips_non_ogr_providers(self):
        """Test that postgres/wms layers are not touched."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <layer-tree-layer providerKey="postgres" source="service='qonnectra'"/>
            <layer-tree-layer providerKey="wms" source="https://wms.example.com"/>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        rewritten = _rewrite_file_datasources(root, ["survey.dxf"], "my-project")

        layers = root.findall(".//layer-tree-layer")
        assert layers[0].get("source") == "service='qonnectra'"
        assert layers[1].get("source") == "https://wms.example.com"
        assert rewritten == []

    def test_skips_gpkg_schema_sources(self):
        """Test that ./schema.gpkg sources (handled by GeoPackage conversion) are skipped."""
        qgs = b"""<?xml version="1.0"?>
        <qgis>
            <maplayer>
                <provider>ogr</provider>
                <datasource>./schema.gpkg|layername=trench</datasource>
            </maplayer>
        </qgis>
        """
        import xml.etree.ElementTree as ET

        root = ET.fromstring(qgs)
        rewritten = _rewrite_file_datasources(root, ["schema.gpkg"], "my-project")

        ds = root.find(".//datasource")
        assert ds.text == "./schema.gpkg|layername=trench"
        assert rewritten == []


class TestQGISDataFileStorage:
    """Tests for QGISDataFileStorage backend."""

    def test_local_dev_location(self, monkeypatch):
        """Test storage uses deployment/qgis/data in local dev."""
        monkeypatch.setattr("apps.api.storage.os.path.exists", lambda p: p != "/app")
        monkeypatch.setattr("apps.api.storage.os.path.isdir", lambda p: False)

        from apps.api.storage import QGISDataFileStorage

        storage = QGISDataFileStorage()
        assert storage.location.endswith("deployment/qgis/data")

    def test_overwrites_existing_file(self, tmp_path):
        """Test that uploading same filename overwrites."""
        from apps.api.storage import QGISDataFileStorage

        storage = QGISDataFileStorage(location=str(tmp_path))
        # Create a subdirectory and file
        sub = tmp_path / "myproject"
        sub.mkdir()
        (sub / "test.dxf").write_text("old content")

        name = storage.get_available_name("myproject/test.dxf")
        assert name == "myproject/test.dxf"
        assert not (sub / "test.dxf").exists()


@pytest.mark.django_db
class TestBuildInquiryExportZip:
    """Tests for build_inquiry_export_zip() service function."""

    @pytest.fixture
    def export_data(self, db):
        """Build a full object graph for export testing.

        Creates all feature types with populated FK fields inside a
        PipelineInquiryArea polygon so they appear in the export.
        """
        import datetime

        from django.contrib.gis.geos import LineString, Point, Polygon

        from apps.api.models import (
            Cable,
            Microduct,
            MicroductCableConnection,
        )

        project = ProjectFactory(project="Test Project")
        flag = FlagFactory(flag="Test Flag")
        status = StatusFactory(status="Active")
        phase = PhaseFactory(phase="Phase 1")
        owner = CompanyFactory(company="Owner Co")
        constructor = CompanyFactory(company="Constructor Co")
        manufacturer = CompanyFactory(company="Manufacturer Co")
        network_level = NetworkLevelFactory(network_level="NE3")
        surface = SurfaceFactory(surface="Asphalt")
        from .factories import ConstructionTypeFactory

        construction_type = ConstructionTypeFactory(construction_type="Open Cut")
        node_type = NodeTypeFactory(node_type="MFG")
        status_dev = StatusDevelopmentFactory(status="Developed")
        md_status = MicroductStatusFactory(microduct_status="Occupied")
        conduit_type = ConduitTypeFactory(conduit_type="Speedpipe 7x10")
        cable_type = CableTypeFactory(
            cable_type="LWL 12F", fiber_count=12, bundle_count=2
        )

        pipeline_record = PipelineRecordFactory(project=project)
        inquiry_area = PipelineInquiryAreaFactory(
            pipeline_record=pipeline_record,
            geom=Polygon(
                ((0, 0), (200, 0), (200, 200), (0, 200), (0, 0)), srid=25832
            ),
        )

        trench1 = TrenchFactory(
            id_trench="TR-AAAAAAA",
            project=project,
            flag=flag,
            surface=surface,
            construction_type=construction_type,
            status=status,
            phase=phase,
            owner=owner,
            constructor=constructor,
            length=100.0,
            date=datetime.date(2025, 6, 15),
            comment="Main route",
            house_connection=False,
            geom=LineString((0, 0), (100, 0), srid=25832),
        )
        trench2 = TrenchFactory(
            id_trench="TR-BBBBBBB",
            project=project,
            flag=flag,
            surface=surface,
            construction_type=construction_type,
            length=50.0,
            geom=LineString((100, 0), (150, 0), srid=25832),
        )
        # Trench outside the inquiry area — should be excluded
        trench_outside = TrenchFactory(
            id_trench="TR-ZZZZZZZ",
            project=project,
            flag=flag,
            geom=LineString((500, 500), (600, 500), srid=25832),
        )

        node = NodeFactory(
            name="MFG-001",
            node_type=node_type,
            project=project,
            flag=flag,
            status=status,
            network_level=network_level,
            owner=owner,
            constructor=constructor,
            manufacturer=manufacturer,
            date=datetime.date(2025, 3, 1),
            geom=Point(50, 50, srid=25832),
        )

        address = AddressFactory(
            id_address="ADR0001",
            zip_code="12345",
            city="Berlin",
            district="Mitte",
            street="Hauptstr.",
            housenumber=42,
            house_number_suffix="a",
            status_development=status_dev,
            project=project,
            flag=flag,
            geom=Point(80, 80, srid=25832),
        )

        area = AreaFactory(
            name="Baugebiet Nord",
            project=project,
            flag=flag,
            geom=Polygon(
                ((10, 10), (90, 10), (90, 90), (10, 90), (10, 10)), srid=25832
            ),
        )

        conduit = ConduitFactory.build(
            name="R01",
            conduit_type=conduit_type,
            outer_conduit="PE50",
            status=status,
            network_level=network_level,
            owner=owner,
            constructor=constructor,
            manufacturer=manufacturer,
            date=datetime.date(2025, 4, 10),
            project=project,
            flag=flag,
        )
        from apps.api.models import Conduit as ConduitModel

        conduit = ConduitModel.objects.create(
            name=conduit.name,
            conduit_type=conduit_type,
            outer_conduit="PE50",
            status=status,
            network_level=network_level,
            owner=owner,
            constructor=constructor,
            manufacturer=manufacturer,
            date=datetime.date(2025, 4, 10),
            project=project,
            flag=flag,
        )

        TrenchConduitConnectionFactory(uuid_trench=trench1, uuid_conduit=conduit)
        TrenchConduitConnectionFactory(uuid_trench=trench2, uuid_conduit=conduit)

        microduct = Microduct.objects.create(
            uuid_conduit=conduit,
            number=1,
            color="rot",
            microduct_status=md_status,
        )

        cable = Cable.objects.create(
            name="LWL-001",
            cable_type=cable_type,
            status=status,
            network_level=network_level,
            owner=owner,
            constructor=constructor,
            manufacturer=manufacturer,
            date=datetime.date(2025, 5, 20),
            uuid_node_start=node,
            uuid_node_end=node,
            length=150.0,
            project=project,
            flag=flag,
        )

        MicroductCableConnection.objects.create(
            uuid_microduct=microduct,
            uuid_cable=cable,
        )
        cable.refresh_from_db()

        return {
            "pipeline_record": pipeline_record,
            "inquiry_area": inquiry_area,
            "trench1": trench1,
            "trench2": trench2,
            "trench_outside": trench_outside,
            "node": node,
            "address": address,
            "area": area,
            "conduit": conduit,
            "microduct": microduct,
            "cable": cable,
            "status": status,
            "phase": phase,
            "owner": owner,
            "constructor": constructor,
            "manufacturer": manufacturer,
            "network_level": network_level,
            "surface": surface,
            "construction_type": construction_type,
            "node_type": node_type,
            "status_dev": status_dev,
            "md_status": md_status,
            "conduit_type": conduit_type,
            "cable_type": cable_type,
            "project": project,
            "flag": flag,
        }

    def _get_zip(self, export_data):
        """Helper to build the export ZIP and return a ZipFile object."""
        buf = build_inquiry_export_zip(export_data["pipeline_record"].uuid)
        return zipfile.ZipFile(buf, "r")

    def _get_layer(self, zf, layer_name):
        """Helper to read and parse a GeoJSON layer from the ZIP."""
        import json

        content = zf.read(f"layers/{layer_name}.geojson")
        return json.loads(content)

    def _get_first_feature(self, zf, layer_name):
        """Helper to get the first feature's properties from a layer."""
        geojson = self._get_layer(zf, layer_name)
        return geojson["features"][0]

    # --- Structure tests ---

    def test_raises_for_missing_inquiry_areas(self, db):
        """ValueError is raised when no inquiry areas exist for the record."""
        record = PipelineRecordFactory()
        with pytest.raises(ValueError, match="No inquiry areas"):
            build_inquiry_export_zip(record.uuid)

    def test_zip_contains_all_layer_files(self, export_data):
        """ZIP contains GeoJSON files for all 7 layers."""
        zf = self._get_zip(export_data)
        names = zf.namelist()
        for layer in [
            "trenches",
            "nodes",
            "addresses",
            "conduits",
            "cables",
            "areas",
            "microducts",
        ]:
            assert f"layers/{layer}.geojson" in names, f"Missing {layer}.geojson"

    def test_zip_contains_qlr_file(self, export_data):
        """ZIP contains a QLR file at the root."""
        zf = self._get_zip(export_data)
        assert "inquiry_export.qlr" in zf.namelist()

    def test_features_outside_area_excluded(self, export_data):
        """Trenches outside the inquiry polygon are not included."""
        zf = self._get_zip(export_data)
        geojson = self._get_layer(zf, "trenches")
        trench_ids = [f["properties"]["id_trench"] for f in geojson["features"]]
        assert "TR-ZZZZZZZ" not in trench_ids
        assert "TR-AAAAAAA" in trench_ids
        assert "TR-BBBBBBB" in trench_ids

    # --- Property tests ---

    def test_trench_properties(self, export_data):
        """Trench GeoJSON features include all FK-resolved labels."""
        zf = self._get_zip(export_data)
        geojson = self._get_layer(zf, "trenches")
        tr1 = next(
            f for f in geojson["features"] if f["properties"]["id_trench"] == "TR-AAAAAAA"
        )
        props = tr1["properties"]

        assert props["uuid"] == str(export_data["trench1"].uuid)
        assert props["surface"] == "Asphalt"
        assert props["construction_type"] == "Open Cut"
        assert props["status"] == "Active"
        assert props["phase"] == "Phase 1"
        assert props["owner"] == "Owner Co"
        assert props["constructor"] == "Constructor Co"
        assert props["length"] == 100.0
        assert props["date"] == "2025-06-15"
        assert props["comment"] == "Main route"
        assert props["house_connection"] is False
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"

    def test_node_properties(self, export_data):
        """Node GeoJSON features include all FK-resolved labels."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "nodes")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["node"].uuid)
        assert props["name"] == "MFG-001"
        assert props["node_type"] == "MFG"
        assert props["status"] == "Active"
        assert props["network_level"] == "NE3"
        assert props["owner"] == "Owner Co"
        assert props["constructor"] == "Constructor Co"
        assert props["manufacturer"] == "Manufacturer Co"
        assert props["date"] == "2025-03-01"
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"

    def test_address_properties(self, export_data):
        """Address GeoJSON features include formatted address and FK labels."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "addresses")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["address"].uuid)
        assert props["id_address"] == "ADR0001"
        assert props["address"] == "Hauptstr. 42a, 12345 Berlin"
        assert props["district"] == "Mitte"
        assert props["status_development"] == "Developed"
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"

    def test_conduit_properties(self, export_data):
        """Conduit GeoJSON features include FK labels and trench_ids."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "conduits")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["conduit"].uuid)
        assert props["name"] == "R01"
        assert props["conduit_type"] == "Speedpipe 7x10"
        assert props["outer_conduit"] == "PE50"
        assert props["status"] == "Active"
        assert props["network_level"] == "NE3"
        assert props["owner"] == "Owner Co"
        assert props["constructor"] == "Constructor Co"
        assert props["manufacturer"] == "Manufacturer Co"
        assert props["date"] == "2025-04-10"
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"
        assert sorted(props["trench_ids"]) == ["TR-AAAAAAA", "TR-BBBBBBB"]

    def test_cable_properties(self, export_data):
        """Cable GeoJSON features include FK labels and conduit_names."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "cables")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["cable"].uuid)
        assert props["name"] == "LWL-001"
        assert props["cable_type"] == "LWL 12F"
        assert props["status"] == "Active"
        assert props["network_level"] == "NE3"
        assert props["owner"] == "Owner Co"
        assert props["constructor"] == "Constructor Co"
        assert props["manufacturer"] == "Manufacturer Co"
        assert props["date"] == "2025-05-20"
        assert props["node_start"] == "MFG-001"
        assert props["node_end"] == "MFG-001"
        assert props["length"] == 150.0
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"
        assert props["conduit_names"] == ["R01"]

    def test_area_properties(self, export_data):
        """Area GeoJSON features include area_type label."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "areas")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["area"].uuid)
        assert props["name"] == "Baugebiet Nord"
        assert "area_type" in props
        assert props["project"] == "Test Project"
        assert props["flag"] == "Test Flag"

    def test_microduct_properties(self, export_data):
        """Microduct GeoJSON features include conduit_name and trench_ids."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "microducts")
        props = feat["properties"]

        assert props["uuid"] == str(export_data["microduct"].uuid)
        assert props["conduit_name"] == "R01"
        assert props["number"] == 1
        assert props["color"] == "rot"
        assert props["microduct_status"] == "Occupied"
        assert sorted(props["trench_ids"]) == ["TR-AAAAAAA", "TR-BBBBBBB"]

    # --- Geometry tests ---

    def test_conduit_geometry_from_trenches(self, export_data):
        """Conduit features have MultiLineString geometry from connected trenches."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "conduits")
        geom = feat["geometry"]

        assert geom is not None
        assert geom["type"] == "MultiLineString"
        assert len(geom["coordinates"]) == 2

    def test_cable_geometry_from_trenches(self, export_data):
        """Cable features have MultiLineString geometry from the trench chain."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "cables")
        geom = feat["geometry"]

        assert geom is not None
        assert geom["type"] == "MultiLineString"
        assert len(geom["coordinates"]) == 2

    def test_microduct_geometry_from_trenches(self, export_data):
        """Microduct features have MultiLineString geometry from parent conduit's trenches."""
        zf = self._get_zip(export_data)
        feat = self._get_first_feature(zf, "microducts")
        geom = feat["geometry"]

        assert geom is not None
        assert geom["type"] == "MultiLineString"
        assert len(geom["coordinates"]) == 2

    # --- Nullable FK test ---

    def test_nullable_fk_renders_as_none(self, db):
        """Null FK fields produce null values in GeoJSON, not errors."""
        from django.contrib.gis.geos import LineString, Polygon

        project = ProjectFactory()
        flag = FlagFactory()
        record = PipelineRecordFactory(project=project)
        PipelineInquiryAreaFactory(
            pipeline_record=record,
            geom=Polygon(
                ((0, 0), (200, 0), (200, 200), (0, 200), (0, 0)), srid=25832
            ),
        )
        TrenchFactory(
            project=project,
            flag=flag,
            status=None,
            phase=None,
            owner=None,
            constructor=None,
            geom=LineString((0, 0), (100, 0), srid=25832),
        )

        buf = build_inquiry_export_zip(record.uuid)
        zf = zipfile.ZipFile(buf, "r")
        geojson = self._get_layer(zf, "trenches")
        props = geojson["features"][0]["properties"]

        assert props["status"] is None
        assert props["phase"] is None
        assert props["owner"] is None
        assert props["constructor"] is None

    # --- QLR tests ---

    def test_qlr_valid_xml_with_layers(self, export_data):
        """QLR file is valid XML with expected layer references and CRS."""
        import xml.etree.ElementTree as ET

        from django.conf import settings

        zf = self._get_zip(export_data)
        qlr_content = zf.read("inquiry_export.qlr").decode("utf-8")
        root = ET.fromstring(qlr_content)

        assert root.tag == "qlr"

        maplayers = root.findall(".//maplayer")
        assert len(maplayers) == 7

        datasources = [ml.findtext("datasource") for ml in maplayers]
        for layer in [
            "trenches",
            "nodes",
            "addresses",
            "conduits",
            "cables",
            "areas",
            "microducts",
        ]:
            assert any(
                f"./layers/{layer}.geojson" in ds for ds in datasources
            ), f"QLR missing datasource for {layer}"

        srid = settings.DEFAULT_SRID
        expected_authid = f"EPSG:{srid}"

        crs_entries = root.findall(".//authid")
        assert all(e.text == expected_authid for e in crs_entries)

        srid_entries = root.findall(".//srid")
        assert len(srid_entries) == 7
        assert all(e.text == str(srid) for e in srid_entries)

        proj4_entries = root.findall(".//proj4")
        assert len(proj4_entries) == 7
        assert all(e.text for e in proj4_entries)

    def test_qlr_has_layer_groups(self, export_data):
        """QLR file has conduits and cables as layer groups."""
        import xml.etree.ElementTree as ET

        zf = self._get_zip(export_data)
        qlr_content = zf.read("inquiry_export.qlr").decode("utf-8")
        root = ET.fromstring(qlr_content)

        groups = root.findall(".//layer-tree-group[@name]")
        group_names = [g.get("name") for g in groups if g.get("name")]
        assert "conduits" in group_names
        assert "cables" in group_names
