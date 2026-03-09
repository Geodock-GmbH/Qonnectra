import json
import logging
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from io import BytesIO

import geopandas as gpd
import openpyxl
import pandas as pd
from django.conf import settings
from django.contrib.contenttypes.models import ContentType
from django.db import connection, transaction
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _
from openpyxl import load_workbook
from pathvalidate import sanitize_filename
from shapely import is_valid, make_valid
from shapely.geometry import LineString, MultiLineString, Point, Polygon, shape, mapping
from shapely.ops import linemerge

from .models import (
    AttributesCompany,
    AttributesConduitType,
    AttributesNetworkLevel,
    AttributesStatus,
    Conduit,
    ConduitTypeColorMapping,
    FeatureFiles,
    Flags,
    Microduct,
    Projects,
    StoragePreferences,
)
from .storage import LocalMediaStorage

logger = logging.getLogger(__name__)


def import_conduits_from_excel(file, max_file_size=10 * 1024 * 1024):
    """
    Imports conduits from an Excel file, validates data, and creates new records.

    Args:
        file: The uploaded Excel file object
        max_file_size: Maximum allowed file size in bytes (default 10MB)

    Returns:
        dict: Contains 'success' boolean and either 'created_count' or 'errors' list
    """
    # File size validation
    file.seek(0, 2)  # Seek to end
    file_size = file.tell()
    file.seek(0)  # Seek back to start

    if file_size > max_file_size:
        return {
            "success": False,
            "errors": [
                str(
                    _("File too large. Maximum size is %(size)dMB.")
                    % {"size": max_file_size // (1024 * 1024)}
                )
            ],
        }

    try:
        workbook = load_workbook(file)
    except Exception as e:
        logger.error(f"Failed to load Excel workbook: {e}")
        return {
            "success": False,
            "errors": [str(_("Invalid Excel file. Could not read workbook."))],
        }

    sheet = workbook.active

    errors = []
    warnings = []
    conduits_to_create = []

    # Get translated headers to match the file
    headers_translated = [
        str(h)
        for h in [
            _("Name"),
            _("Type"),
            _("Outer Conduit"),
            _("Status"),
            _("Network Level"),
            _("Owner"),
            _("Constructor"),
            _("Manufacturer"),
            _("Date"),
            _("Project"),
            _("Flag"),
        ]
    ]

    header_from_file = [cell.value for cell in sheet[1]]

    # Validate headers exist
    if not header_from_file or all(h is None for h in header_from_file):
        return {
            "success": False,
            "errors": [
                str(
                    _(
                        "No headers found in Excel file. First row must contain column headers."
                    )
                )
            ],
        }

    # Map file headers to a more usable format
    header_map = {
        headers_translated[0]: "name",
        headers_translated[1]: "conduit_type",
        headers_translated[2]: "outer_conduit",
        headers_translated[3]: "status",
        headers_translated[4]: "network_level",
        headers_translated[5]: "owner",
        headers_translated[6]: "constructor",
        headers_translated[7]: "manufacturer",
        headers_translated[8]: "date",
        headers_translated[9]: "project",
        headers_translated[10]: "flag",
    }

    # Check for required 'Name' header
    if headers_translated[0] not in header_from_file:
        return {
            "success": False,
            "errors": [
                str(
                    _("Required 'Name' column not found. Expected header: '%(header)s'")
                    % {"header": headers_translated[0]}
                )
            ],
        }

    # Check for unrecognized columns (warning, not error)
    unmapped_headers = [h for h in header_from_file if h and h not in header_map]
    if unmapped_headers:
        warnings.append(
            str(
                _("Unrecognized columns will be ignored: %(columns)s")
                % {"columns": ", ".join(unmapped_headers)}
            )
        )

    mapped_header = [header_map.get(h) for h in header_from_file]

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        row_data = dict(zip(mapped_header, row))

        name = row_data.get("name")
        if not name:
            errors.append(_("Row %(row)d: Name is required.") % {"row": row_idx})
            continue

        if Conduit.objects.filter(name=name).exists():
            errors.append(
                _("Row %(row)d: Conduit with name '%(name)s' already exists.")
                % {"row": row_idx, "name": name}
            )
            continue

        try:
            project_val = row_data.get("project")
            project = Projects.objects.get(project=project_val) if project_val else None

            flag_val = row_data.get("flag")
            flag = Flags.objects.get(flag=flag_val) if flag_val else None

            conduit_type_val = row_data.get("conduit_type")
            conduit_type = (
                AttributesConduitType.objects.get(conduit_type=conduit_type_val)
                if conduit_type_val
                else None
            )

            status_val = row_data.get("status")
            status = (
                AttributesStatus.objects.get(status=status_val) if status_val else None
            )

            network_level_val = row_data.get("network_level")
            network_level = (
                AttributesNetworkLevel.objects.get(network_level=network_level_val)
                if network_level_val
                else None
            )

            # Handle company lookups individually for better error reporting
            owner_val = row_data.get("owner")
            owner = None
            if owner_val:
                try:
                    owner = AttributesCompany.objects.get(company=owner_val)
                except AttributesCompany.DoesNotExist:
                    errors.append(
                        f'Row {row_idx}: Owner company "{owner_val}" not found.'
                    )
                    continue

            constructor_val = row_data.get("constructor")
            constructor = None
            if constructor_val:
                try:
                    constructor = AttributesCompany.objects.get(company=constructor_val)
                except AttributesCompany.DoesNotExist:
                    errors.append(
                        f'Row {row_idx}: Constructor company "{constructor_val}" not found.'
                    )
                    continue

            manufacturer_val = row_data.get("manufacturer")
            manufacturer = None
            if manufacturer_val:
                try:
                    manufacturer = AttributesCompany.objects.get(
                        company=manufacturer_val
                    )
                except AttributesCompany.DoesNotExist:
                    errors.append(
                        f'Row {row_idx}: Manufacturer company "{manufacturer_val}" not found.'
                    )
                    continue

            conduit = Conduit(
                name=name,
                conduit_type=conduit_type,
                outer_conduit=row_data.get("outer_conduit"),
                status=status,
                network_level=network_level,
                owner=owner,
                constructor=constructor,
                manufacturer=manufacturer,
                date=row_data.get("date"),
                project=project,
                flag=flag,
            )
            conduits_to_create.append(conduit)

        except Projects.DoesNotExist:
            errors.append(f'Row {row_idx}: Project "{project_val}" not found.')
        except Flags.DoesNotExist:
            errors.append(f'Row {row_idx}: Flag "{flag_val}" not found.')
        except AttributesConduitType.DoesNotExist:
            errors.append(
                f'Row {row_idx}: Conduit Type "{conduit_type_val}" not found.'
            )
        except AttributesStatus.DoesNotExist:
            errors.append(f'Row {row_idx}: Status "{status_val}" not found.')
        except AttributesNetworkLevel.DoesNotExist:
            errors.append(
                f'Row {row_idx}: Network Level "{network_level_val}" not found.'
            )

        except Exception as e:
            errors.append(f"Row {row_idx}: An unexpected error occurred: {e}")

    if errors:
        return {"success": False, "errors": errors, "warnings": warnings}

    try:
        with transaction.atomic():
            created_conduits = Conduit.objects.bulk_create(conduits_to_create)

            # Manually create microducts since bulk_create bypasses post_save signals
            for conduit in created_conduits:
                if conduit.conduit_type:
                    color_mappings = (
                        ConduitTypeColorMapping.objects.filter(
                            conduit_type=conduit.conduit_type
                        )
                        .select_related("color")
                        .order_by("position")
                    )
                    for mapping in color_mappings:
                        Microduct.objects.create(
                            uuid_conduit=conduit,
                            number=mapping.position,
                            color=mapping.color.name_de,
                        )
    except Exception as e:
        logger.error(f"Failed to bulk create conduits: {e}")
        return {
            "success": False,
            "errors": [
                str(_("Failed to save to database: %(error)s") % {"error": str(e)})
            ],
        }

    result = {"success": True, "created_count": len(conduits_to_create)}
    if warnings:
        result["warnings"] = warnings
    return result


def generate_conduit_import_template():
    """
    Creates a basic Excel workbook and returns it as an HTTP response.
    """
    # Create a workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Conduit Import Template"

    headers = [
        str(h)
        for h in [
            _("Name"),
            _("Type"),
            _("Outer Conduit"),
            _("Status"),
            _("Network Level"),
            _("Owner"),
            _("Constructor"),
            _("Manufacturer"),
            _("Date"),
            _("Project"),
            _("Flag"),
        ]
    ]
    for col, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col, value=header)

    # Add one example at row 2
    example_row = [
        "RV1.1.1",
        "12x10/6",
        "",
        "geplant",
        "Hausanschluss-Ebene",
        "Geodock",
        "Geodock",
        "Geodock",
        "2025-01-01",
        "Default",
        "Default",
    ]

    for col, value in enumerate(example_row, start=1):
        worksheet.cell(row=2, column=col, value=value)

    # Prepare the response for Excel format
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="conduit_import_template.xlsx"'
    )

    # Save the workbook to the response
    workbook.save(response)

    return response


def rename_feature_folder(instance, old_identifier, new_identifier):
    """
    Rename storage folder when a feature's identifier changes.

    Updates both the filesystem folder and all FeatureFiles.file_path entries
    in the database within an atomic transaction.

    Args:
        instance: Model instance (Node, Cable, Conduit, Trench, Address, or ResidentialUnit)
        old_identifier: Previous identifier value (name, id_trench, address string, or id_residential_unit)
        new_identifier: New identifier value

    Raises:
        OSError: If folder rename fails (rolls back the entire operation)
    """
    prefs = StoragePreferences.objects.first()
    model_name = instance._meta.model_name

    # ResidentialUnit derives project from parent address
    if model_name == "residentialunit":
        address = instance.uuid_address
        project_name = (
            address.project.project
            if hasattr(address, "project") and address.project
            else "default"
        )
    else:
        project_name = (
            instance.project.project
            if hasattr(instance, "project") and instance.project
            else "default"
        )
    project_name = sanitize_filename(project_name)

    old_feature_folder = sanitize_filename(str(old_identifier))
    new_feature_folder = sanitize_filename(str(new_identifier))

    # ResidentialUnit folders are nested under the parent address folder
    if model_name == "residentialunit":
        suffix = address.house_number_suffix or ""
        address_id = sanitize_filename(
            f"{address.street} {address.housenumber}{suffix}, "
            f"{address.zip_code} {address.city}"
        )

        if prefs and prefs.folder_structure:
            address_folder_config = prefs.folder_structure.get("address", {})
            address_base = address_folder_config.get("default", "addresses")
            ru_folder_config = prefs.folder_structure.get("residentialunit", {})
            ru_base = ru_folder_config.get("default", "residential_units")
            if "/" in ru_base:
                ru_base = ru_base.split("/", 1)[0]
        else:
            address_base = "addresses"
            ru_base = "residential_units"

        old_path = (
            f"{project_name}/{address_base}/{address_id}/{ru_base}/{old_feature_folder}"
        )
        new_path = (
            f"{project_name}/{address_base}/{address_id}/{ru_base}/{new_feature_folder}"
        )
    else:
        if prefs and prefs.folder_structure:
            folder_config = prefs.folder_structure.get(model_name, {})
            if isinstance(folder_config, dict):
                for category, path in folder_config.items():
                    if "/" in path:
                        base_folder = path.split("/", 1)[0]
                        break
                    else:
                        base_folder = path
                        break
                else:
                    base_folder = f"{model_name}s"
            else:
                base_folder = folder_config if folder_config else f"{model_name}s"
        else:
            base_folder = f"{model_name}s"

        old_path = f"{project_name}/{base_folder}/{old_feature_folder}"
        new_path = f"{project_name}/{base_folder}/{new_feature_folder}"

    storage = LocalMediaStorage()
    content_type = ContentType.objects.get_for_model(instance)

    with transaction.atomic():
        folder_existed = storage.rename_folder(old_path, new_path)

        if folder_existed:
            files = FeatureFiles.objects.select_for_update().filter(
                content_type=content_type, object_id=instance.pk
            )
            updated_count = 0
            for f in files:
                old_file_path = f.file_path.name
                if old_path in old_file_path:
                    f.file_path.name = old_file_path.replace(old_path, new_path, 1)
                    f.save(update_fields=["file_path"])
                    updated_count += 1

            logger.info(
                f"Renamed feature folder for {model_name} "
                f"'{old_identifier}' -> '{new_identifier}', "
                f"updated {updated_count} file path(s)"
            )
        else:
            logger.debug(
                f"No folder to rename for {model_name} "
                f"'{old_identifier}' (no files uploaded yet)"
            )


def move_file_to_feature(file_obj, target_feature, target_content_type):
    """
    Move a FeatureFile to a different feature, updating both the database
    and the physical file location.

    Args:
        file_obj: FeatureFiles instance to move
        target_feature: Target model instance (Node, Cable, Conduit, Trench, Address, ResidentialUnit, or Area)
        target_content_type: ContentType for the target model

    Returns:
        tuple: (success: bool, new_path: str or None, error: str or None)
    """
    storage = LocalMediaStorage()
    old_path = file_obj.file_path.name

    prefs = StoragePreferences.objects.first()
    model_name = target_content_type.model

    if model_name == "trench":
        feature_id = target_feature.id_trench
    elif model_name in ("conduit", "cable", "node", "area"):
        feature_id = target_feature.name
    elif model_name == "address":
        suffix = target_feature.house_number_suffix or ""
        feature_id = (
            f"{target_feature.street} {target_feature.housenumber}{suffix}, "
            f"{target_feature.zip_code} {target_feature.city}"
        )
    elif model_name == "residentialunit":
        feature_id = target_feature.id_residential_unit or str(target_feature.pk)
    else:
        feature_id = str(target_feature.pk)

    feature_id = sanitize_filename(str(feature_id))

    # ResidentialUnit derives project from parent address
    if model_name == "residentialunit":
        address = target_feature.uuid_address
        project_name = (
            address.project.project
            if hasattr(address, "project") and address.project
            else "default"
        )
    else:
        project_name = (
            target_feature.project.project
            if hasattr(target_feature, "project") and target_feature.project
            else "default"
        )
    project_name = sanitize_filename(project_name)

    file_extension = file_obj.file_type or ""
    file_extension = file_extension.lower()

    try:
        from .models import FileTypeCategory

        category_obj = FileTypeCategory.objects.get(extension=file_extension)
        file_category = category_obj.category
    except FileTypeCategory.DoesNotExist:
        file_category = "documents"

    if (
        model_name == "residentialunit"
        and prefs
        and prefs.mode == "AUTO"
        and prefs.folder_structure
    ):
        # ResidentialUnit uses nested path under parent address folder
        address = target_feature.uuid_address
        suffix = address.house_number_suffix or ""
        address_id = sanitize_filename(
            f"{address.street} {address.housenumber}{suffix}, "
            f"{address.zip_code} {address.city}"
        )
        address_folder_paths = prefs.folder_structure.get("address", {})
        address_folder_name = address_folder_paths.get("default", "addresses")
        ru_folder_paths = prefs.folder_structure.get("residentialunit", {})
        ru_folder_name = ru_folder_paths.get(
            file_category, ru_folder_paths.get("default", "residential_units")
        )
        if "/" in ru_folder_name:
            ru_base, sub_folder = ru_folder_name.split("/", 1)
            new_path = (
                f"{project_name}/{address_folder_name}/{address_id}/"
                f"{ru_base}/{feature_id}/{sub_folder}/"
                f"{file_obj.file_name}.{file_obj.file_type}"
            )
        else:
            new_path = (
                f"{project_name}/{address_folder_name}/{address_id}/"
                f"{ru_folder_name}/{feature_id}/"
                f"{file_obj.file_name}.{file_obj.file_type}"
            )
    elif model_name == "residentialunit":
        # Fallback without preferences
        address = target_feature.uuid_address
        suffix = address.house_number_suffix or ""
        address_id = sanitize_filename(
            f"{address.street} {address.housenumber}{suffix}, "
            f"{address.zip_code} {address.city}"
        )
        new_path = (
            f"{project_name}/addresses/{address_id}/residential_units/{feature_id}/"
            f"{file_obj.file_name}.{file_obj.file_type}"
        )
    elif prefs and prefs.mode == "AUTO" and prefs.folder_structure:
        folder_paths = prefs.folder_structure.get(model_name, {})
        folder_name = folder_paths.get(
            file_category, folder_paths.get("default", model_name + "s")
        )

        if "/" in folder_name:
            base_folder, sub_folder = folder_name.split("/", 1)
            new_path = (
                f"{project_name}/{base_folder}/{feature_id}/{sub_folder}/"
                f"{file_obj.file_name}.{file_obj.file_type}"
            )
        else:
            new_path = (
                f"{project_name}/{folder_name}/{feature_id}/"
                f"{file_obj.file_name}.{file_obj.file_type}"
            )
    else:
        new_path = (
            f"{project_name}/{model_name}s/{feature_id}/"
            f"{file_obj.file_name}.{file_obj.file_type}"
        )

    if storage.exists(new_path):
        return (False, None, f"File already exists at target path: {new_path}")

    try:
        if storage.exists(old_path):
            old_file = storage.open(old_path, "rb")
            content = old_file.read()
            old_file.close()

            from django.core.files.base import ContentFile

            storage.save(new_path, ContentFile(content))

            storage.delete(old_path)

            logger.info(f"Moved file from {old_path} to {new_path}")
        else:
            logger.warning(
                f"Physical file not found at {old_path}, updating database path only"
            )
    except Exception as e:
        logger.error(f"Error moving file from {old_path} to {new_path}: {e}")
        return (False, None, str(e))

    return (True, new_path, None)


# Layer configuration registry mapping layer names to database metadata
GEOPACKAGE_LAYER_CONFIG = {
    # Geometry layers (with geom field)
    "trench": {
        "db_table": "trench",
        "geometry_type": "LineString",
        "pk": "uuid",
        "geom_column": "geom",
    },
    "node": {
        "db_table": "node",
        "geometry_type": "Point",
        "pk": "uuid",
        "geom_column": "geom",
    },
    "address": {
        "db_table": "address",
        "geometry_type": "Point",
        "pk": "uuid",
        "geom_column": "geom",
    },
    "area": {
        "db_table": "area",
        "geometry_type": "Polygon",
        "pk": "uuid",
        "geom_column": "geom",
    },
    # Non-geometry layers (attribute/relation tables)
    "conduit": {
        "db_table": "conduit",
        "geometry_type": None,
        "pk": "uuid",
    },
    "microduct": {
        "db_table": "microduct",
        "geometry_type": None,
        "pk": "uuid",
    },
    "cable": {
        "db_table": "cable",
        "geometry_type": None,
        "pk": "uuid",
    },
    # Attribute lookup tables
    "attributes_surface": {
        "db_table": "attributes_surface",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_construction_type": {
        "db_table": "attributes_construction_type",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_status": {
        "db_table": "attributes_status",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_phase": {
        "db_table": "attributes_phase",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_company": {
        "db_table": "attributes_company",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_node_type": {
        "db_table": "attributes_node_type",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_conduit_type": {
        "db_table": "attributes_conduit_type",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_network_level": {
        "db_table": "attributes_network_level",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_status_development": {
        "db_table": "attributes_status_development",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_microduct_status": {
        "db_table": "attributes_microduct_status",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_cable_type": {
        "db_table": "attributes_cable_type",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_microduct_color": {
        "db_table": "attributes_microduct_color",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_fiber_status": {
        "db_table": "attributes_fiber_status",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_fiber_color": {
        "db_table": "attributes_fiber_color",
        "geometry_type": None,
        "pk": "id",
    },
    "attributes_area_type": {
        "db_table": "attributes_area_type",
        "geometry_type": None,
        "pk": "id",
    },
    "projects": {
        "db_table": "projects",
        "geometry_type": None,
        "pk": "id",
    },
    "flags": {
        "db_table": "flags",
        "geometry_type": None,
        "pk": "id",
    },
}


def _get_table_columns(table_name: str) -> list[dict]:
    """
    Get column information for a database table.

    Returns list of dicts with 'name', 'type', 'nullable' keys.
    """
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT column_name, data_type, is_nullable, udt_name
            FROM information_schema.columns
            WHERE table_name = %s AND table_schema = 'public'
            ORDER BY ordinal_position
            """,
            [table_name],
        )
        columns = []
        for row in cursor.fetchall():
            columns.append(
                {
                    "name": row[0],
                    "type": row[1],
                    "nullable": row[2] == "YES",
                    "udt_name": row[3],
                }
            )
        return columns


def _postgres_type_to_pandas(pg_type: str, udt_name: str) -> str:
    """Map PostgreSQL data type to pandas dtype string."""
    type_mapping = {
        "uuid": "object",
        "character varying": "object",
        "text": "object",
        "integer": "Int64",
        "bigint": "Int64",
        "smallint": "Int64",
        "numeric": "float64",
        "double precision": "float64",
        "real": "float64",
        "boolean": "boolean",
        "date": "object",
        "timestamp without time zone": "object",
        "timestamp with time zone": "object",
        "json": "object",
        "jsonb": "object",
        "USER-DEFINED": "object",  # geometry, etc.
    }
    return type_mapping.get(pg_type, "object")


def generate_geopackage_schema(layers: list[str] | None = None) -> HttpResponse:
    """
    Generate an empty GeoPackage file with the database schema.

    Args:
        layers: Optional list of layer names to include. If None, includes all layers.

    Returns:
        HttpResponse with the GeoPackage file as attachment.
    """
    if layers is None:
        layers = list(GEOPACKAGE_LAYER_CONFIG.keys())
    else:
        # Validate layer names
        invalid = [layer for layer in layers if layer not in GEOPACKAGE_LAYER_CONFIG]
        if invalid:
            raise ValueError(f"Invalid layer names: {invalid}")

    srid = getattr(settings, "DEFAULT_SRID", 25832)

    with tempfile.NamedTemporaryFile(suffix=".gpkg", delete=False) as tmp_file:
        gpkg_path = tmp_file.name

    try:
        for layer_name in layers:
            config = GEOPACKAGE_LAYER_CONFIG[layer_name]
            table_name = config["db_table"]
            geom_type = config.get("geometry_type")
            geom_column = config.get("geom_column", "geom")

            columns = _get_table_columns(table_name)
            if not columns:
                logger.warning(f"No columns found for table {table_name}, skipping")
                continue

            data = {}
            for col in columns:
                if col["udt_name"] == "geometry":
                    continue
                dtype = _postgres_type_to_pandas(col["type"], col["udt_name"])
                data[col["name"]] = pd.Series(dtype=dtype)

            df = pd.DataFrame(data)

            if geom_type:
                geometry_class = {
                    "Point": Point,
                    "LineString": LineString,
                    "Polygon": Polygon,
                }
                geom_cls = geometry_class.get(geom_type)
                if geom_cls:
                    gdf = gpd.GeoDataFrame(
                        df,
                        geometry=gpd.GeoSeries([], crs=f"EPSG:{srid}"),
                        crs=f"EPSG:{srid}",
                    )
                    gdf = gdf.rename_geometry(geom_column)
                else:
                    gdf = gpd.GeoDataFrame(df, crs=f"EPSG:{srid}")
            else:
                gdf = gpd.GeoDataFrame(df)

            gdf.to_file(gpkg_path, driver="GPKG", layer=layer_name)
            logger.info(f"Added layer '{layer_name}' to GeoPackage")

        with open(gpkg_path, "rb") as f:
            gpkg_content = f.read()

        response = HttpResponse(
            gpkg_content, content_type="application/geopackage+sqlite3"
        )
        response["Content-Disposition"] = 'attachment; filename="schema.gpkg"'
        return response

    finally:
        import os

        if os.path.exists(gpkg_path):
            os.unlink(gpkg_path)


def _build_postgres_datasource(
    layer_name: str, pg_service: str, srid: int
) -> str | None:
    """
    Build PostgreSQL datasource string for a layer.

    Returns None if layer is not in configuration.
    """
    config = GEOPACKAGE_LAYER_CONFIG.get(layer_name)
    if not config:
        return None

    table_name = config["db_table"]
    pk = config["pk"]
    geom_type = config.get("geometry_type")
    geom_column = config.get("geom_column", "geom")

    if geom_type:
        return (
            f"service='{pg_service}' key='{pk}' srid={srid} type={geom_type} "
            f'checkPrimaryKeyUnicity=\'0\' table="public"."{table_name}" ({geom_column})'
        )
    else:
        return (
            f"service='{pg_service}' key='{pk}' "
            f'checkPrimaryKeyUnicity=\'0\' table="public"."{table_name}"'
        )


def _extract_layer_name_from_gpkg_source(source: str) -> str | None:
    """
    Extract layer name from GeoPackage datasource string.

    Example: "./schema.gpkg|layername=address" -> "address"
    """
    if "|layername=" in source:
        return source.split("|layername=")[-1]
    return None


def convert_qgs_to_postgres(qgs_content: bytes) -> bytes:
    """
    Transform QGS XML datasources from GeoPackage to PostgreSQL pg_service.

    Args:
        qgs_content: Raw bytes of QGS XML file.

    Returns:
        Transformed QGS XML content as bytes.
    """
    pg_service = getattr(settings, "QGIS_PG_SERVICE_NAME", "qonnectra")
    srid = getattr(settings, "DEFAULT_SRID", 25832)

    tree = ET.ElementTree(ET.fromstring(qgs_content))
    root = tree.getroot()

    converted_layers = []

    for layer_tree in root.iter("layer-tree-layer"):
        provider_key = layer_tree.get("providerKey")
        source = layer_tree.get("source", "")

        if provider_key == "ogr" and "|layername=" in source:
            layer_name = _extract_layer_name_from_gpkg_source(source)
            if layer_name:
                pg_source = _build_postgres_datasource(layer_name, pg_service, srid)
                if pg_source:
                    layer_tree.set("providerKey", "postgres")
                    layer_tree.set("source", pg_source)
                    converted_layers.append(layer_name)

    for maplayer in root.iter("maplayer"):
        provider_elem = maplayer.find("provider")
        datasource_elem = maplayer.find("datasource")

        if provider_elem is not None and datasource_elem is not None:
            if provider_elem.text == "ogr" and datasource_elem.text:
                layer_name = _extract_layer_name_from_gpkg_source(datasource_elem.text)
                if layer_name:
                    pg_source = _build_postgres_datasource(layer_name, pg_service, srid)
                    if pg_source:
                        provider_elem.text = "postgres"
                        datasource_elem.text = pg_source

    for gps_settings in root.iter("ProjectGpsSettings"):
        provider = gps_settings.get("destinationLayerProvider")
        source = gps_settings.get("destinationLayerSource", "")

        if provider == "ogr" and "|layername=" in source:
            layer_name = _extract_layer_name_from_gpkg_source(source)
            if layer_name:
                pg_source = _build_postgres_datasource(layer_name, pg_service, srid)
                if pg_source:
                    gps_settings.set("destinationLayerProvider", "postgres")
                    gps_settings.set("destinationLayerSource", pg_source)

    logger.info(
        f"Converted {len(converted_layers)} layers to PostgreSQL: {converted_layers}"
    )

    output = BytesIO()
    tree.write(output, encoding="utf-8", xml_declaration=True)
    return output.getvalue()


def handle_qgis_file(file_content: bytes, filename: str) -> tuple[bytes, bool]:
    """
    Extract QGS content from QGS or QGZ file.

    Args:
        file_content: Raw file bytes
        filename: Original filename to determine format

    Returns:
        Tuple of (qgs_content_bytes, is_qgz)
    """
    if filename.lower().endswith(".qgz"):
        with zipfile.ZipFile(BytesIO(file_content), "r") as zf:
            qgs_files = [n for n in zf.namelist() if n.endswith(".qgs")]
            if not qgs_files:
                raise ValueError("No .qgs file found inside .qgz archive")
            qgs_content = zf.read(qgs_files[0])
            return qgs_content, True
    else:
        return file_content, False


def repackage_qgz(
    qgs_content: bytes, original_qgz: bytes, qgs_filename: str = None
) -> bytes:
    """
    Repackage modified QGS content back into QGZ format.

    Args:
        qgs_content: Modified QGS XML bytes
        original_qgz: Original QGZ file bytes (to preserve other files)
        qgs_filename: Optional QGS filename to use inside the archive

    Returns:
        New QGZ file bytes
    """
    output = BytesIO()

    with zipfile.ZipFile(BytesIO(original_qgz), "r") as original_zip:
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as new_zip:
            for item in original_zip.namelist():
                if item.endswith(".qgs"):
                    new_zip.writestr(item, qgs_content)
                    if qgs_filename is None:
                        qgs_filename = item
                else:
                    new_zip.writestr(item, original_zip.read(item))

    return output.getvalue()


# =============================================================================
# Fiber Trace Service
# =============================================================================


def trace_fiber(
    fiber_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Trace a single fiber through all its splice connections bidirectionally.
    Uses PostgreSQL recursive CTE for performance at scale (1M+ splices).
    Returns a flat list of trace segments that can be assembled into a tree.
    Includes:
    - Address info for nodes
    - Residential unit info for connected endpoints
    - Cable endpoint nodes (uuid_node_start/uuid_node_end) for full path visibility

    Args:
        fiber_id: UUID of the fiber to trace
        include_geometry: If True, include trench geometry
        geometry_mode: "segments" for individual trenches, "merged" for combined
        orient_geometry: If True, orient geometries from cable start to end
    """
    sql = """
    WITH RECURSIVE container_hierarchy AS (
        -- Base case: top-level containers (no parent)
        SELECT
            c.uuid as container_id,
            c.uuid_node,
            c.parent_container,
            c.name as container_name,
            ct.name as container_type_name,
            ARRAY[jsonb_build_object(
                'type', ct.name,
                'name', c.name
            )] as path,
            1 as depth
        FROM container c
        JOIN container_type ct ON ct.id = c.container_type
        WHERE c.parent_container IS NULL

        UNION ALL

        -- Recursive case: children containers
        SELECT
            c.uuid as container_id,
            c.uuid_node,
            c.parent_container,
            c.name as container_name,
            ct.name as container_type_name,
            ch.path || jsonb_build_object(
                'type', ct.name,
                'name', c.name
            ),
            ch.depth + 1
        FROM container c
        JOIN container_type ct ON ct.id = c.container_type
        JOIN container_hierarchy ch ON ch.container_id = c.parent_container
        WHERE ch.depth < 10
    ),
    fiber_trace AS (
        -- Base case: starting fiber
        SELECT
            f.uuid as fiber_id,
            f.uuid_cable as cable_id,
            f.fiber_number_absolute,
            f.bundle_number,
            f.fiber_color,
            afc_fiber.hex_code as fiber_color_hex,
            f.fiber_number_in_bundle,
            f.bundle_color,
            afc_bundle.hex_code as bundle_color_hex,
            f.layer,
            fs_status.fiber_status as fiber_status,
            ct.cable_type as cable_type_name,
            c.name as cable_name,
            NULL::uuid as from_splice_id,
            NULL::uuid as from_node_id,
            NULL::text as from_node_name,
            NULL::text as direction,
            NULL::uuid as prev_fiber_id,
            0 as depth,
            ARRAY[f.uuid] as visited
        FROM fiber f
        JOIN cable c ON c.uuid = f.uuid_cable
        LEFT JOIN attributes_fiber_status fs_status ON fs_status.id = f.fiber_status
        LEFT JOIN attributes_cable_type ct ON ct.id = c.cable_type
        LEFT JOIN attributes_fiber_color afc_fiber ON LOWER(afc_fiber.name_de) = LOWER(f.fiber_color)
        LEFT JOIN attributes_fiber_color afc_bundle ON LOWER(afc_bundle.name_de) = LOWER(f.bundle_color)
        WHERE f.uuid = %(fiber_id)s

        UNION ALL

        -- Recursive case: follow splices
        SELECT
            next_fiber.uuid as fiber_id,
            next_fiber.uuid_cable as cable_id,
            next_fiber.fiber_number_absolute,
            next_fiber.bundle_number,
            next_fiber.fiber_color,
            next_afc_fiber.hex_code as fiber_color_hex,
            next_fiber.fiber_number_in_bundle,
            next_fiber.bundle_color,
            next_afc_bundle.hex_code as bundle_color_hex,
            next_fiber.layer,
            next_fs_status.fiber_status as fiber_status,
            next_ct.cable_type as cable_type_name,
            next_cable.name as cable_name,
            fs.uuid as from_splice_id,
            n.uuid as from_node_id,
            n.name as from_node_name,
            CASE
                WHEN fs.fiber_a = ft.fiber_id OR fs.shared_fiber_a = ft.fiber_id THEN 'a_to_b'
                ELSE 'b_to_a'
            END as direction,
            ft.fiber_id as prev_fiber_id,
            ft.depth + 1,
            ft.visited || next_fiber.uuid
        FROM fiber_trace ft
        JOIN fiber_splice fs ON (
            fs.fiber_a = ft.fiber_id OR
            fs.fiber_b = ft.fiber_id OR
            fs.shared_fiber_a = ft.fiber_id OR
            fs.shared_fiber_b = ft.fiber_id
        )
        JOIN node_structure ns ON ns.uuid = fs.node_structure
        JOIN node n ON n.uuid = ns.uuid_node
        LEFT JOIN fiber next_fiber ON next_fiber.uuid = (
            CASE
                WHEN fs.fiber_a = ft.fiber_id THEN COALESCE(fs.fiber_b, fs.shared_fiber_b)
                WHEN fs.fiber_b = ft.fiber_id THEN COALESCE(fs.fiber_a, fs.shared_fiber_a)
                WHEN fs.shared_fiber_a = ft.fiber_id THEN COALESCE(fs.fiber_b, fs.shared_fiber_b)
                WHEN fs.shared_fiber_b = ft.fiber_id THEN COALESCE(fs.fiber_a, fs.shared_fiber_a)
            END
        )
        LEFT JOIN cable next_cable ON next_cable.uuid = next_fiber.uuid_cable
        LEFT JOIN attributes_fiber_status next_fs_status ON next_fs_status.id = next_fiber.fiber_status
        LEFT JOIN attributes_cable_type next_ct ON next_ct.id = next_cable.cable_type
        LEFT JOIN attributes_fiber_color next_afc_fiber ON LOWER(next_afc_fiber.name_de) = LOWER(next_fiber.fiber_color)
        LEFT JOIN attributes_fiber_color next_afc_bundle ON LOWER(next_afc_bundle.name_de) = LOWER(next_fiber.bundle_color)
        WHERE next_fiber.uuid IS NOT NULL
          AND NOT (next_fiber.uuid = ANY(ft.visited))
          AND ft.depth < 100
    )
    SELECT
        ft.fiber_id,
        ft.cable_id,
        ft.fiber_number_absolute,
        ft.bundle_number,
        ft.fiber_color,
        ft.fiber_color_hex,
        ft.fiber_number_in_bundle,
        ft.bundle_color,
        ft.bundle_color_hex,
        ft.layer,
        ft.fiber_status,
        ft.cable_type_name,
        ft.cable_name,
        ft.from_splice_id,
        ft.from_node_id,
        ft.from_node_name,
        ft.direction,
        ft.depth,
        -- Full address fields from node
        addr.uuid as address_id,
        addr.id_address as address_id_address,
        addr.street as address_street,
        addr.housenumber as address_housenumber,
        addr.house_number_suffix as address_suffix,
        addr.zip_code as address_zip_code,
        addr.city as address_city,
        addr.district as address_district,
        addr_status.status as address_status_development,
        addr_project.project as address_project,
        addr_flag.flag as address_flag,
        -- Residential units connected to THIS fiber (aggregated as JSON array)
        fiber_ru_lookup.residential_units as connected_residential_units,
        -- Cable endpoint nodes (for full path visibility)
        cable.uuid_node_start as cable_node_start_id,
        cable.uuid_node_end as cable_node_end_id,
        node_start.name as cable_node_start_name,
        node_end.name as cable_node_end_name,
        node_start_type.node_type as cable_node_start_type,
        node_end_type.node_type as cable_node_end_type,
        -- Start node address
        start_addr.uuid as cable_start_address_id,
        start_addr.street as cable_start_address_street,
        start_addr.housenumber as cable_start_address_housenumber,
        start_addr.house_number_suffix as cable_start_address_suffix,
        start_addr.zip_code as cable_start_address_zip_code,
        start_addr.city as cable_start_address_city,
        -- End node address
        end_addr.uuid as cable_end_address_id,
        end_addr.street as cable_end_address_street,
        end_addr.housenumber as cable_end_address_housenumber,
        end_addr.house_number_suffix as cable_end_address_suffix,
        end_addr.zip_code as cable_end_address_zip_code,
        end_addr.city as cable_end_address_city,
        -- Splice component info
        ft.from_splice_id as splice_id,
        fs_splice.port_number as splice_port_number,
        comp_type.component_type as component_type_name,
        comp_struct.in_or_out as component_in_or_out,
        comp_struct.port as component_structure_port,
        comp_struct.port_alias as component_structure_port_alias,
        ns.slot_start as component_slot_start,
        ns.slot_end as component_slot_end,
        nsc.side as component_slot_side,
        nsc.container as component_container_id,
        -- Container path (from slot configuration's container)
        ch.path as container_path
    FROM fiber_trace ft
    LEFT JOIN node n ON n.uuid = ft.from_node_id
    LEFT JOIN address addr ON addr.uuid = n.uuid_address
    LEFT JOIN attributes_status_development addr_status
        ON addr_status.id = addr.status_development
    LEFT JOIN projects addr_project ON addr_project.id = addr.project
    LEFT JOIN flags addr_flag ON addr_flag.id = addr.flag
    -- Find RU connected to THIS fiber via any splice
    LEFT JOIN LATERAL (
        SELECT jsonb_agg(
            jsonb_build_object(
                'id', ru.uuid,
                'id_residential_unit', ru.id_residential_unit,
                'floor', ru.floor,
                'side', ru.side,
                'building_section', ru.building_section,
                'type', ru_type.residential_unit_type,
                'status', ru_status.status,
                'external_id_1', ru.external_id_1,
                'external_id_2', ru.external_id_2,
                'resident_name', ru.resident_name,
                'resident_recorded_date', ru.resident_recorded_date,
                'ready_for_service', ru.ready_for_service,
                'address_id', ru_addr.uuid,
                'address_street', ru_addr.street,
                'address_housenumber', ru_addr.housenumber,
                'address_suffix', ru_addr.house_number_suffix,
                'address_zip_code', ru_addr.zip_code,
                'address_city', ru_addr.city
            )
        ) as residential_units
        FROM fiber_splice fs2
        LEFT JOIN residential_unit ru ON ru.uuid = CASE
            WHEN fs2.fiber_a = ft.fiber_id OR fs2.shared_fiber_a = ft.fiber_id
            THEN fs2.residential_unit_b_id
            WHEN fs2.fiber_b = ft.fiber_id OR fs2.shared_fiber_b = ft.fiber_id
            THEN fs2.residential_unit_a_id
        END
        LEFT JOIN attributes_residential_unit_type ru_type ON ru_type.id = ru.residential_unit_type
        LEFT JOIN attributes_residential_unit_status ru_status ON ru_status.id = ru.status
        LEFT JOIN address ru_addr ON ru_addr.uuid = ru.uuid_address
        WHERE (fs2.fiber_a = ft.fiber_id OR fs2.fiber_b = ft.fiber_id
               OR fs2.shared_fiber_a = ft.fiber_id OR fs2.shared_fiber_b = ft.fiber_id)
          AND ru.uuid IS NOT NULL
    ) fiber_ru_lookup ON true
    -- Join cable endpoint nodes
    LEFT JOIN cable ON cable.uuid = ft.cable_id
    LEFT JOIN node node_start ON node_start.uuid = cable.uuid_node_start
    LEFT JOIN node node_end ON node_end.uuid = cable.uuid_node_end
    LEFT JOIN attributes_node_type node_start_type ON node_start_type.id = node_start.node_type
    LEFT JOIN attributes_node_type node_end_type ON node_end_type.id = node_end.node_type
    LEFT JOIN address start_addr ON start_addr.uuid = node_start.uuid_address
    LEFT JOIN address end_addr ON end_addr.uuid = node_end.uuid_address
    -- Component info from splice
    LEFT JOIN fiber_splice fs_splice ON fs_splice.uuid = ft.from_splice_id
    LEFT JOIN node_structure ns ON ns.uuid = fs_splice.node_structure
    LEFT JOIN attributes_component_type comp_type ON comp_type.id = ns.component_type
    LEFT JOIN attributes_component_structure comp_struct ON comp_struct.id = ns.component_structure
    LEFT JOIN node_slot_configuration nsc ON nsc.uuid = ns.slot_configuration
    LEFT JOIN container_hierarchy ch ON ch.container_id = nsc.container
    ORDER BY ft.depth;
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"fiber_id": str(fiber_id)})
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    return _build_trace_result(
        rows, "fiber", fiber_id, include_geometry, geometry_mode, orient_geometry
    )


def trace_cable(
    cable_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Trace all fibers in a cable through their splice connections.
    """
    sql = """
    SELECT uuid FROM fiber WHERE uuid_cable = %(cable_id)s
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"cable_id": str(cable_id)})
        fiber_ids = [row[0] for row in cursor.fetchall()]

    if not fiber_ids:
        return {
            "entry_point": {"type": "cable", "id": str(cable_id)},
            "trace_trees": [],
            "cable_infrastructure": {},
            "statistics": {
                "total_fibers": 0,
                "total_nodes": 0,
                "total_splices": 0,
                "total_addresses": 0,
                "total_residential_units": 0,
                "total_cables": 0,
                "total_trenches": 0,
                "has_branches": False,
            },
        }

    all_traces = []
    for fiber_id in fiber_ids:
        trace = trace_fiber(fiber_id, include_geometry, geometry_mode, orient_geometry)
        all_traces.append(trace)

    total_fibers = sum(t["statistics"]["total_fibers"] for t in all_traces)
    total_nodes = sum(t["statistics"]["total_nodes"] for t in all_traces)
    total_splices = sum(t["statistics"]["total_splices"] for t in all_traces)
    total_addresses = sum(t["statistics"]["total_addresses"] for t in all_traces)
    total_rus = sum(t["statistics"]["total_residential_units"] for t in all_traces)
    total_cables = sum(t["statistics"]["total_cables"] for t in all_traces)
    total_trenches = sum(t["statistics"]["total_trenches"] for t in all_traces)
    has_branches = any(t["statistics"]["has_branches"] for t in all_traces)

    # Merge cable_infrastructure from all traces
    merged_infrastructure = {}
    for trace in all_traces:
        for cable_id_str, infra in trace.get("cable_infrastructure", {}).items():
            if cable_id_str not in merged_infrastructure:
                merged_infrastructure[cable_id_str] = infra

    return {
        "entry_point": {"type": "cable", "id": str(cable_id)},
        "trace_trees": [t["trace_tree"] for t in all_traces if t["trace_tree"]],
        "cable_infrastructure": merged_infrastructure,
        "statistics": {
            "total_fibers": total_fibers,
            "total_nodes": total_nodes,
            "total_splices": total_splices,
            "total_addresses": total_addresses,
            "total_residential_units": total_rus,
            "total_cables": total_cables,
            "total_trenches": total_trenches,
            "has_branches": has_branches,
        },
    }


def trace_node(
    node_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Trace all fibers passing through a node.
    Includes fibers from:
    1. Splices at this node (via node_structure)
    2. Cables that start or end at this node (uuid_node_start/uuid_node_end)
    """
    sql = """
    SELECT DISTINCT fiber_id FROM (
        -- Fibers through splices at this node
        SELECT DISTINCT
            COALESCE(fs.fiber_a, fs.shared_fiber_a, fs.fiber_b, fs.shared_fiber_b) as fiber_id
        FROM fiber_splice fs
        JOIN node_structure ns ON ns.uuid = fs.node_structure
        WHERE ns.uuid_node = %(node_id)s
          AND (fs.fiber_a IS NOT NULL OR fs.shared_fiber_a IS NOT NULL
               OR fs.fiber_b IS NOT NULL OR fs.shared_fiber_b IS NOT NULL)

        UNION

        -- Fibers from cables that start at this node
        SELECT f.uuid as fiber_id
        FROM cable c
        JOIN fiber f ON f.uuid_cable = c.uuid
        WHERE c.uuid_node_start = %(node_id)s

        UNION

        -- Fibers from cables that end at this node
        SELECT f.uuid as fiber_id
        FROM cable c
        JOIN fiber f ON f.uuid_cable = c.uuid
        WHERE c.uuid_node_end = %(node_id)s
    ) combined
    WHERE fiber_id IS NOT NULL
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"node_id": str(node_id)})
        fiber_ids = [row[0] for row in cursor.fetchall() if row[0]]

    if not fiber_ids:
        return {
            "entry_point": {"type": "node", "id": str(node_id)},
            "trace_trees": [],
            "cable_infrastructure": {},
            "statistics": {
                "total_fibers": 0,
                "total_nodes": 0,
                "total_splices": 0,
                "total_addresses": 0,
                "total_residential_units": 0,
                "total_cables": 0,
                "total_trenches": 0,
                "has_branches": False,
            },
        }

    all_traces = []
    seen_fibers = set()

    for fiber_id in fiber_ids:
        if fiber_id in seen_fibers:
            continue
        trace = trace_fiber(fiber_id, include_geometry, geometry_mode, orient_geometry)
        for segment in trace.get("_raw_segments", []):
            seen_fibers.add(segment["fiber_id"])
        all_traces.append(trace)

    total_fibers = sum(t["statistics"]["total_fibers"] for t in all_traces)
    total_nodes = sum(t["statistics"]["total_nodes"] for t in all_traces)
    total_splices = sum(t["statistics"]["total_splices"] for t in all_traces)
    total_addresses = sum(t["statistics"]["total_addresses"] for t in all_traces)
    total_rus = sum(t["statistics"]["total_residential_units"] for t in all_traces)
    total_cables = sum(t["statistics"]["total_cables"] for t in all_traces)
    total_trenches = sum(t["statistics"]["total_trenches"] for t in all_traces)
    has_branches = any(t["statistics"]["has_branches"] for t in all_traces)

    # Merge cable_infrastructure from all traces
    merged_infrastructure = {}
    for trace in all_traces:
        for cable_id_str, infra in trace.get("cable_infrastructure", {}).items():
            if cable_id_str not in merged_infrastructure:
                merged_infrastructure[cable_id_str] = infra

    return {
        "entry_point": {"type": "node", "id": str(node_id)},
        "trace_trees": [t["trace_tree"] for t in all_traces if t["trace_tree"]],
        "cable_infrastructure": merged_infrastructure,
        "statistics": {
            "total_fibers": total_fibers,
            "total_nodes": total_nodes,
            "total_splices": total_splices,
            "total_addresses": total_addresses,
            "total_residential_units": total_rus,
            "total_cables": total_cables,
            "total_trenches": total_trenches,
            "has_branches": has_branches,
        },
    }


def trace_address(
    address_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Trace all fibers connected to an address via:
    1. Nodes linked to this address (node.uuid_address)
    2. Residential units under this address that have fiber splices
    """
    sql = """
    SELECT DISTINCT fiber_id FROM (
        -- Fibers through nodes linked to this address
        SELECT DISTINCT
            COALESCE(fs.fiber_a, fs.shared_fiber_a, fs.fiber_b, fs.shared_fiber_b) as fiber_id
        FROM node n
        JOIN node_structure ns ON ns.uuid_node = n.uuid
        JOIN fiber_splice fs ON fs.node_structure = ns.uuid
        WHERE n.uuid_address = %(address_id)s
          AND (fs.fiber_a IS NOT NULL OR fs.shared_fiber_a IS NOT NULL
               OR fs.fiber_b IS NOT NULL OR fs.shared_fiber_b IS NOT NULL)

        UNION

        -- Fibers connected to residential units under this address
        SELECT DISTINCT
            COALESCE(fs.fiber_a, fs.shared_fiber_a, fs.fiber_b, fs.shared_fiber_b) as fiber_id
        FROM residential_unit ru
        JOIN fiber_splice fs ON (
            fs.residential_unit_a_id = ru.uuid OR fs.residential_unit_b_id = ru.uuid
        )
        WHERE ru.uuid_address = %(address_id)s
          AND (fs.fiber_a IS NOT NULL OR fs.shared_fiber_a IS NOT NULL
               OR fs.fiber_b IS NOT NULL OR fs.shared_fiber_b IS NOT NULL)
    ) combined
    WHERE fiber_id IS NOT NULL
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"address_id": str(address_id)})
        fiber_ids = [row[0] for row in cursor.fetchall()]

    if not fiber_ids:
        return {
            "entry_point": {"type": "address", "id": str(address_id)},
            "trace_trees": [],
            "cable_infrastructure": {},
            "statistics": {
                "total_fibers": 0,
                "total_nodes": 0,
                "total_splices": 0,
                "total_addresses": 0,
                "total_residential_units": 0,
                "total_cables": 0,
                "total_trenches": 0,
                "has_branches": False,
            },
        }

    all_traces = []
    seen_fibers = set()

    for fiber_id in fiber_ids:
        if fiber_id in seen_fibers:
            continue
        trace = trace_fiber(fiber_id, include_geometry, geometry_mode, orient_geometry)
        for segment in trace.get("_raw_segments", []):
            seen_fibers.add(segment["fiber_id"])
        all_traces.append(trace)

    total_fibers = sum(t["statistics"]["total_fibers"] for t in all_traces)
    total_nodes = sum(t["statistics"]["total_nodes"] for t in all_traces)
    total_splices = sum(t["statistics"]["total_splices"] for t in all_traces)
    total_addresses = sum(t["statistics"]["total_addresses"] for t in all_traces)
    total_rus = sum(t["statistics"]["total_residential_units"] for t in all_traces)
    total_cables = sum(t["statistics"]["total_cables"] for t in all_traces)
    total_trenches = sum(t["statistics"]["total_trenches"] for t in all_traces)
    has_branches = any(t["statistics"]["has_branches"] for t in all_traces)

    # Merge cable_infrastructure from all traces
    merged_infrastructure = {}
    for trace in all_traces:
        for cable_id_str, infra in trace.get("cable_infrastructure", {}).items():
            if cable_id_str not in merged_infrastructure:
                merged_infrastructure[cable_id_str] = infra

    return {
        "entry_point": {"type": "address", "id": str(address_id)},
        "trace_trees": [t["trace_tree"] for t in all_traces if t["trace_tree"]],
        "cable_infrastructure": merged_infrastructure,
        "statistics": {
            "total_fibers": total_fibers,
            "total_nodes": total_nodes,
            "total_splices": total_splices,
            "total_addresses": total_addresses,
            "total_residential_units": total_rus,
            "total_cables": total_cables,
            "total_trenches": total_trenches,
            "has_branches": has_branches,
        },
    }


def trace_residential_unit(
    residential_unit_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Trace all fibers connected to a residential unit via fiber_splice.
    """
    sql = """
    SELECT DISTINCT
        COALESCE(fs.fiber_a, fs.shared_fiber_a, fs.fiber_b, fs.shared_fiber_b) as fiber_id
    FROM fiber_splice fs
    WHERE (fs.residential_unit_a_id = %(ru_id)s OR fs.residential_unit_b_id = %(ru_id)s)
      AND (fs.fiber_a IS NOT NULL OR fs.shared_fiber_a IS NOT NULL
           OR fs.fiber_b IS NOT NULL OR fs.shared_fiber_b IS NOT NULL)
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"ru_id": str(residential_unit_id)})
        fiber_ids = [row[0] for row in cursor.fetchall() if row[0]]

    if not fiber_ids:
        return {
            "entry_point": {"type": "residential_unit", "id": str(residential_unit_id)},
            "trace_trees": [],
            "cable_infrastructure": {},
            "statistics": {
                "total_fibers": 0,
                "total_nodes": 0,
                "total_splices": 0,
                "total_addresses": 0,
                "total_residential_units": 0,
                "total_cables": 0,
                "total_trenches": 0,
                "has_branches": False,
            },
        }

    all_traces = []
    seen_fibers = set()

    for fiber_id in fiber_ids:
        if fiber_id in seen_fibers:
            continue
        trace = trace_fiber(fiber_id, include_geometry, geometry_mode, orient_geometry)
        for segment in trace.get("_raw_segments", []):
            seen_fibers.add(segment["fiber_id"])
        all_traces.append(trace)

    total_fibers = sum(t["statistics"]["total_fibers"] for t in all_traces)
    total_nodes = sum(t["statistics"]["total_nodes"] for t in all_traces)
    total_splices = sum(t["statistics"]["total_splices"] for t in all_traces)
    total_addresses = sum(t["statistics"]["total_addresses"] for t in all_traces)
    total_rus = sum(t["statistics"]["total_residential_units"] for t in all_traces)
    total_cables = sum(t["statistics"]["total_cables"] for t in all_traces)
    total_trenches = sum(t["statistics"]["total_trenches"] for t in all_traces)
    has_branches = any(t["statistics"]["has_branches"] for t in all_traces)

    # Merge cable_infrastructure from all traces
    merged_infrastructure = {}
    for trace in all_traces:
        for cable_id_str, infra in trace.get("cable_infrastructure", {}).items():
            if cable_id_str not in merged_infrastructure:
                merged_infrastructure[cable_id_str] = infra

    return {
        "entry_point": {"type": "residential_unit", "id": str(residential_unit_id)},
        "trace_trees": [t["trace_tree"] for t in all_traces if t["trace_tree"]],
        "cable_infrastructure": merged_infrastructure,
        "statistics": {
            "total_fibers": total_fibers,
            "total_nodes": total_nodes,
            "total_splices": total_splices,
            "total_addresses": total_addresses,
            "total_residential_units": total_rus,
            "total_cables": total_cables,
            "total_trenches": total_trenches,
            "has_branches": has_branches,
        },
    }


def _build_trace_result(
    rows: list,
    entry_type: str,
    entry_id,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
) -> dict:
    """
    Build the trace result structure from flat database rows.
    Includes address, residential unit, and cable endpoint information.

    Args:
        rows: Flat database rows from trace query
        entry_type: Type of entry point (fiber, cable, node, etc.)
        entry_id: UUID of entry point
        include_geometry: If True, include trench geometry
        geometry_mode: "segments" for individual trenches, "merged" for combined
        orient_geometry: If True, orient geometries from cable start to end
    """
    if not rows:
        return {
            "entry_point": {"type": entry_type, "id": str(entry_id)},
            "trace_tree": None,
            "statistics": {
                "total_fibers": 0,
                "total_nodes": 0,
                "total_splices": 0,
                "total_addresses": 0,
                "total_residential_units": 0,
                "has_branches": False,
            },
            "_raw_segments": [],
        }

    nodes_seen = set()
    splices_seen = set()
    fibers_seen = set()
    addresses_seen = set()
    residential_units_seen = set()
    cable_endpoints_seen = {}

    for row in rows:
        fibers_seen.add(row["fiber_id"])
        if row["from_node_id"]:
            nodes_seen.add(row["from_node_id"])
        if row["from_splice_id"]:
            splices_seen.add(row["from_splice_id"])
        if row.get("address_id"):
            addresses_seen.add(row["address_id"])
        connected_rus = row.get("connected_residential_units") or []
        if isinstance(connected_rus, str):
            connected_rus = json.loads(connected_rus)
        for ru in connected_rus:
            if ru and isinstance(ru, dict) and ru.get("id"):
                residential_units_seen.add(ru["id"])
            if ru and isinstance(ru, dict) and ru.get("address_id"):
                addresses_seen.add(ru["address_id"])
        # Track cable endpoint nodes
        cable_id = row["cable_id"]
        if cable_id not in cable_endpoints_seen:
            cable_endpoints_seen[cable_id] = {
                "cable_id": str(cable_id),
                "cable_name": row["cable_name"],
                "start_node": None,
                "end_node": None,
            }
            if row.get("cable_node_start_id"):
                nodes_seen.add(row["cable_node_start_id"])
                cable_endpoints_seen[cable_id]["start_node"] = {
                    "id": str(row["cable_node_start_id"]),
                    "name": row.get("cable_node_start_name"),
                    "type": row.get("cable_node_start_type"),
                }
                if row.get("cable_start_address_id"):
                    addresses_seen.add(row["cable_start_address_id"])
                    cable_endpoints_seen[cable_id]["start_node"]["address"] = {
                        "id": str(row["cable_start_address_id"]),
                        "street": row.get("cable_start_address_street"),
                        "housenumber": row.get("cable_start_address_housenumber"),
                        "suffix": row.get("cable_start_address_suffix") or "",
                        "zip_code": row.get("cable_start_address_zip_code"),
                        "city": row.get("cable_start_address_city"),
                    }
            if row.get("cable_node_end_id"):
                nodes_seen.add(row["cable_node_end_id"])
                cable_endpoints_seen[cable_id]["end_node"] = {
                    "id": str(row["cable_node_end_id"]),
                    "name": row.get("cable_node_end_name"),
                    "type": row.get("cable_node_end_type"),
                }
                if row.get("cable_end_address_id"):
                    addresses_seen.add(row["cable_end_address_id"])
                    cable_endpoints_seen[cable_id]["end_node"]["address"] = {
                        "id": str(row["cable_end_address_id"]),
                        "street": row.get("cable_end_address_street"),
                        "housenumber": row.get("cable_end_address_housenumber"),
                        "suffix": row.get("cable_end_address_suffix") or "",
                        "zip_code": row.get("cable_end_address_zip_code"),
                        "city": row.get("cable_end_address_city"),
                    }

    def build_node_with_address(row):
        """Build node dict including full address details if available."""
        if not row["from_node_id"]:
            return None

        node_data = {
            "id": str(row["from_node_id"]),
            "name": row["from_node_name"],
        }

        if row.get("address_id"):
            node_data["address"] = {
                "id": str(row["address_id"]),
                "id_address": row.get("address_id_address"),
                "street": row["address_street"],
                "housenumber": row["address_housenumber"],
                "suffix": row.get("address_suffix") or "",
                "zip_code": row["address_zip_code"],
                "city": row["address_city"],
                "district": row.get("address_district"),
                "status_development": row.get("address_status_development"),
                "project": row.get("address_project"),
                "flag": row.get("address_flag"),
            }

        return node_data

    def build_residential_units(row):
        """Build list of residential unit dicts for all connected units."""
        connected_rus = row.get("connected_residential_units") or []
        if isinstance(connected_rus, str):
            connected_rus = json.loads(connected_rus)
        if not connected_rus:
            return None

        result = []
        for ru in connected_rus:
            if not ru or not isinstance(ru, dict) or not ru.get("id"):
                continue

            ru_data = {
                "id": str(ru["id"]),
                "id_residential_unit": ru.get("id_residential_unit"),
                "floor": ru.get("floor"),
                "side": ru.get("side"),
                "building_section": ru.get("building_section"),
                "type": ru.get("type"),
                "status": ru.get("status"),
                "external_id_1": ru.get("external_id_1"),
                "external_id_2": ru.get("external_id_2"),
                "resident_name": ru.get("resident_name"),
                "resident_recorded_date": ru.get("resident_recorded_date"),
                "ready_for_service": ru.get("ready_for_service"),
            }

            if ru.get("address_id"):
                ru_data["address"] = {
                    "id": str(ru["address_id"]),
                    "street": ru.get("address_street"),
                    "housenumber": ru.get("address_housenumber"),
                    "suffix": ru.get("address_suffix") or "",
                    "zip_code": ru.get("address_zip_code"),
                    "city": ru.get("address_city"),
                }

            result.append(ru_data)

        return result if result else None

    def build_cable_endpoints(row):
        """Build cable endpoint info from row data."""
        cable_id = row["cable_id"]
        endpoints = cable_endpoints_seen.get(cable_id)
        if endpoints and (endpoints["start_node"] or endpoints["end_node"]):
            return endpoints
        return None

    def build_splice(row):
        """Build splice dict with component hierarchy."""
        if not row.get("splice_id"):
            return None

        # Parse container path - each element may be a JSON string or already a dict
        container_path = row.get("container_path") or []
        parsed_path = []
        for item in container_path:
            if isinstance(item, str):
                parsed_path.append(json.loads(item))
            else:
                parsed_path.append(item)

        splice_data = {
            "id": str(row["splice_id"]),
            "port_number": row.get("splice_port_number"),
            "component": {
                "type": row.get("component_type_name"),
                "in_or_out": row.get("component_in_or_out"),
                "structure_port": row.get("component_structure_port"),
                "structure_port_alias": row.get("component_structure_port_alias"),
                "slot_start": row.get("component_slot_start"),
                "slot_end": row.get("component_slot_end"),
                "slot_side": row.get("component_slot_side"),
            },
            "container_path": parsed_path,
        }
        return splice_data

    root = rows[0]
    trace_tree = {
        "fiber": {
            "id": str(root["fiber_id"]),
            "fiber_number_absolute": root["fiber_number_absolute"],
            "fiber_number_in_bundle": root.get("fiber_number_in_bundle"),
            "bundle_number": root["bundle_number"],
            "bundle_color": root.get("bundle_color"),
            "bundle_color_hex": root.get("bundle_color_hex"),
            "fiber_color": root["fiber_color"],
            "fiber_color_hex": root.get("fiber_color_hex"),
            "layer": root.get("layer"),
            "status": root.get("fiber_status"),
            "cable_id": str(root["cable_id"]),
            "cable_name": root["cable_name"],
            "cable_type": root.get("cable_type_name"),
        },
        "cable_endpoints": build_cable_endpoints(root),
        "splice": build_splice(root),
        "node": build_node_with_address(root),
        "residential_units": build_residential_units(root),
        "direction": root.get("direction"),
        "children": [],
    }

    nodes_by_depth = {}
    for row in rows[1:]:
        depth = row["depth"]
        if depth not in nodes_by_depth:
            nodes_by_depth[depth] = []
        nodes_by_depth[depth].append(row)

    def build_children(parent_fiber_id, current_depth):
        children = []
        if current_depth not in nodes_by_depth:
            return children

        for row in nodes_by_depth[current_depth]:
            child = {
                "fiber": {
                    "id": str(row["fiber_id"]),
                    "fiber_number_absolute": row["fiber_number_absolute"],
                    "fiber_number_in_bundle": row.get("fiber_number_in_bundle"),
                    "bundle_number": row["bundle_number"],
                    "bundle_color": row.get("bundle_color"),
                    "bundle_color_hex": row.get("bundle_color_hex"),
                    "fiber_color": row["fiber_color"],
                    "fiber_color_hex": row.get("fiber_color_hex"),
                    "layer": row.get("layer"),
                    "status": row.get("fiber_status"),
                    "cable_id": str(row["cable_id"]),
                    "cable_name": row["cable_name"],
                    "cable_type": row.get("cable_type_name"),
                },
                "cable_endpoints": build_cable_endpoints(row),
                "splice": build_splice(row),
                "node": build_node_with_address(row),
                "residential_units": build_residential_units(row),
                "direction": row["direction"],
                "children": build_children(row["fiber_id"], current_depth + 1),
            }
            children.append(child)

        return children

    trace_tree["children"] = build_children(root["fiber_id"], 1)

    has_branches = any(len(nodes_by_depth.get(d, [])) > 1 for d in nodes_by_depth)

    # Collect unique cable IDs
    cables_seen = set()
    for row in rows:
        cables_seen.add(row["cable_id"])

    # Build cable endpoints with geometry for orientation
    cable_endpoints_for_geometry = {}
    if orient_geometry and include_geometry:
        # Fetch node geometries for cable endpoints
        cable_node_ids = set()
        for cable_id, endpoints in cable_endpoints_seen.items():
            if endpoints.get("start_node") and endpoints["start_node"].get("id"):
                cable_node_ids.add(endpoints["start_node"]["id"])
            if endpoints.get("end_node") and endpoints["end_node"].get("id"):
                cable_node_ids.add(endpoints["end_node"]["id"])

        # Query node geometries
        node_geoms = {}
        if cable_node_ids:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT uuid, ST_X(geom) as x, ST_Y(geom) as y
                    FROM node
                    WHERE uuid = ANY(%(node_ids)s)
                    """,
                    {"node_ids": [str(nid) for nid in cable_node_ids]},
                )
                for row in cursor.fetchall():
                    if row[1] is not None and row[2] is not None:
                        node_geoms[str(row[0])] = Point(row[1], row[2])

        # Map cable endpoints to geometry
        for cable_id, endpoints in cable_endpoints_seen.items():
            cable_endpoints_for_geometry[str(cable_id)] = {
                "start_geom": node_geoms.get(endpoints["start_node"]["id"])
                if endpoints.get("start_node")
                else None,
                "end_geom": node_geoms.get(endpoints["end_node"]["id"])
                if endpoints.get("end_node")
                else None,
            }

    # Get infrastructure for all cables
    cable_infrastructure = _get_cable_infrastructure(
        list(cables_seen),
        include_geometry,
        geometry_mode,
        orient_geometry,
        cable_endpoints_for_geometry,
    )

    return {
        "entry_point": {"type": entry_type, "id": str(entry_id)},
        "trace_tree": trace_tree,
        "cable_infrastructure": cable_infrastructure,
        "statistics": {
            "total_fibers": len(fibers_seen),
            "total_nodes": len(nodes_seen),
            "total_splices": len(splices_seen),
            "total_addresses": len(addresses_seen),
            "total_residential_units": len(residential_units_seen),
            "total_cables": len(cables_seen),
            "total_trenches": sum(len(inf["trenches"]) for inf in cable_infrastructure.values()),
            "has_branches": has_branches,
        },
        "_raw_segments": rows,
    }


# =============================================================================
# Geometry Processing Helpers
# =============================================================================


def _validate_and_clean_geometry(geojson_dict: dict) -> dict | None:
    """
    Validate and clean a GeoJSON geometry.

    Args:
        geojson_dict: GeoJSON geometry dict

    Returns:
        Cleaned GeoJSON dict or None if invalid/empty
    """
    if not geojson_dict:
        return None

    try:
        geom = shape(geojson_dict)
        if geom.is_empty:
            return None
        if not is_valid(geom):
            geom = make_valid(geom)
        return mapping(geom)
    except Exception:
        return None


def _merge_trench_geometries(trenches: list[dict]) -> dict | None:
    """
    Merge multiple trench geometries into a single geometry.

    Args:
        trenches: List of trench dicts with 'geometry' field containing GeoJSON

    Returns:
        GeoJSON dict (LineString or MultiLineString depending on connectivity)
        or None if no valid geometries
    """
    if not trenches:
        return None

    shapely_geoms = []
    for trench in trenches:
        geom_json = trench.get("geometry")
        if not geom_json:
            continue

        cleaned = _validate_and_clean_geometry(geom_json)
        if cleaned:
            try:
                shapely_geoms.append(shape(cleaned))
            except Exception:
                continue

    if not shapely_geoms:
        return None

    if len(shapely_geoms) == 1:
        return mapping(shapely_geoms[0])

    # linemerge returns LineString if connected, MultiLineString if disconnected
    merged = linemerge(shapely_geoms)
    return mapping(merged)


def _orient_geometry(
    geom_dict: dict, cable_start_geom: Point | None, cable_end_geom: Point | None
) -> dict:
    """
    Orient a LineString/MultiLineString so it flows from cable start to end.
    Compares first/last points of geometry to cable endpoint nodes.

    Args:
        geom_dict: GeoJSON geometry dict
        cable_start_geom: Shapely Point of cable start node
        cable_end_geom: Shapely Point of cable end node

    Returns:
        Oriented GeoJSON geometry dict
    """
    if not geom_dict or not cable_start_geom:
        return geom_dict

    try:
        geom = shape(geom_dict)
    except Exception:
        return geom_dict

    if geom.is_empty:
        return geom_dict

    if geom.geom_type == "LineString":
        first_pt = Point(geom.coords[0])
        last_pt = Point(geom.coords[-1])

        # If last point is closer to cable start than first point, reverse
        dist_first_to_start = first_pt.distance(cable_start_geom)
        dist_last_to_start = last_pt.distance(cable_start_geom)
        if dist_last_to_start < dist_first_to_start:
            return mapping(geom.reverse())
        return geom_dict

    elif geom.geom_type == "MultiLineString":
        # Orient each component LineString
        oriented_lines = []
        for line in geom.geoms:
            first_pt = Point(line.coords[0])
            last_pt = Point(line.coords[-1])

            dist_first_to_start = first_pt.distance(cable_start_geom)
            dist_last_to_start = last_pt.distance(cable_start_geom)
            if dist_last_to_start < dist_first_to_start:
                oriented_lines.append(line.reverse())
            else:
                oriented_lines.append(line)

        return mapping(MultiLineString(oriented_lines))

    return geom_dict


def _get_cable_infrastructure(
    cable_ids: list,
    include_geometry: bool = False,
    geometry_mode: str = "segments",
    orient_geometry: bool = False,
    cable_endpoints: dict | None = None,
) -> dict:
    """
    Get infrastructure path for cables: microduct → conduit → trenches.

    Args:
        cable_ids: List of cable UUIDs to fetch infrastructure for
        include_geometry: If True, include trench geometry as GeoJSON
        geometry_mode: "segments" for individual trenches, "merged" for combined geometry
        orient_geometry: If True, orient geometries from cable start to end
        cable_endpoints: Dict mapping cable_id to {start_geom, end_geom} Point objects

    Returns:
        Dict mapping cable_id to infrastructure data
    """
    if not cable_ids:
        return {}

    if cable_endpoints is None:
        cable_endpoints = {}

    geometry_select = "ST_AsGeoJSON(t.geom)::jsonb" if include_geometry else "NULL"

    sql = f"""
    SELECT
        c.uuid as cable_id,
        md.uuid as microduct_id,
        md.number as microduct_number,
        md.color as microduct_color,
        amc.hex_code as microduct_color_hex,
        md_status.microduct_status as microduct_status,
        cond.uuid as conduit_id,
        cond.name as conduit_name,
        cond_type.conduit_type as conduit_type,
        t.uuid as trench_id,
        t.id_trench,
        const_type.construction_type,
        surf.surface,
        t.length as trench_length,
        {geometry_select} as trench_geometry
    FROM cable c
    LEFT JOIN microduct_cable_connection mcc ON mcc.uuid_cable = c.uuid
    LEFT JOIN microduct md ON md.uuid = mcc.uuid_microduct
    LEFT JOIN attributes_microduct_status md_status ON md_status.id = md.microduct_status
    LEFT JOIN attributes_microduct_color amc ON LOWER(amc.name_de) = LOWER(md.color)
    LEFT JOIN conduit cond ON cond.uuid = md.uuid_conduit
    LEFT JOIN attributes_conduit_type cond_type ON cond_type.id = cond.conduit_type
    LEFT JOIN trench_conduit_connect tcc ON tcc.uuid_conduit = cond.uuid
    LEFT JOIN trench t ON t.uuid = tcc.uuid_trench
    LEFT JOIN attributes_construction_type const_type ON const_type.id = t.construction_type
    LEFT JOIN attributes_surface surf ON surf.id = t.surface
    WHERE c.uuid = ANY(%(cable_ids)s)
    ORDER BY c.uuid, md.number, t.id_trench
    """

    with connection.cursor() as cursor:
        cursor.execute(sql, {"cable_ids": [str(cid) for cid in cable_ids]})
        columns = [col[0] for col in cursor.description]
        rows = [dict(zip(columns, row)) for row in cursor.fetchall()]

    # Group by cable
    infrastructure = {}
    for row in rows:
        cable_id = str(row["cable_id"])
        if cable_id not in infrastructure:
            infrastructure[cable_id] = {
                "microduct": None,
                "conduit": None,
                "trenches": [],
            }

        # Set microduct (first one wins, should be same for all rows of a cable)
        if row.get("microduct_id") and not infrastructure[cable_id]["microduct"]:
            infrastructure[cable_id]["microduct"] = {
                "id": str(row["microduct_id"]),
                "number": row["microduct_number"],
                "color": row["microduct_color"],
                "color_hex": row.get("microduct_color_hex"),
                "status": row.get("microduct_status"),
            }

        # Set conduit
        if row.get("conduit_id") and not infrastructure[cable_id]["conduit"]:
            infrastructure[cable_id]["conduit"] = {
                "id": str(row["conduit_id"]),
                "name": row["conduit_name"],
                "type": row.get("conduit_type"),
            }

        # Add trench (dedupe by checking if already added)
        if row.get("trench_id"):
            trench_id = str(row["trench_id"])
            existing_ids = [t["id"] for t in infrastructure[cable_id]["trenches"]]
            if trench_id not in existing_ids:
                # Parse geometry if it's a string
                trench_geom = row.get("trench_geometry")
                if isinstance(trench_geom, str):
                    try:
                        trench_geom = json.loads(trench_geom)
                    except (json.JSONDecodeError, TypeError):
                        trench_geom = None

                infrastructure[cable_id]["trenches"].append({
                    "id": trench_id,
                    "id_trench": row["id_trench"],
                    "construction_type": row.get("construction_type"),
                    "surface": row.get("surface"),
                    "length": float(row["trench_length"]) if row.get("trench_length") else None,
                    "geometry": trench_geom,
                })

    # Process geometries based on mode
    if include_geometry:
        for cable_id, infra in infrastructure.items():
            trenches = infra.get("trenches", [])
            endpoints = cable_endpoints.get(cable_id, {})
            start_geom = endpoints.get("start_geom")
            end_geom = endpoints.get("end_geom")

            if geometry_mode == "merged":
                # Merge all trench geometries into one
                merged = _merge_trench_geometries(trenches)
                if merged and orient_geometry:
                    merged = _orient_geometry(merged, start_geom, end_geom)
                infra["merged_geometry"] = merged

                # Calculate total length from merged geometry
                if merged:
                    try:
                        merged_geom = shape(merged)
                        infra["total_length"] = merged_geom.length
                    except Exception:
                        infra["total_length"] = sum(
                            t.get("length") or 0 for t in trenches
                        )

                # Remove individual geometries to reduce payload
                for trench in trenches:
                    trench.pop("geometry", None)

            elif orient_geometry:
                # Segments mode with orientation - orient each trench geometry
                for trench in trenches:
                    if trench.get("geometry"):
                        trench["geometry"] = _orient_geometry(
                            trench["geometry"], start_geom, end_geom
                        )

    return infrastructure


def trace_fiber_summary(fiber_id) -> dict:
    """
    Get a compact trace summary for a fiber.
    Returns start/end endpoints and key statistics.
    Reuses trace_fiber() and extracts summary data.

    Finds the true terminal nodes by collecting all cable endpoints
    and excluding nodes that appear as splice points (middle of the path).
    """
    full_trace = trace_fiber(fiber_id, include_geometry=False)

    # Remove internal data
    if "_raw_segments" in full_trace:
        del full_trace["_raw_segments"]

    tree = full_trace.get("trace_tree")
    stats = full_trace.get("statistics", {})

    # Collect all endpoint nodes and splice nodes to find terminals
    all_endpoint_nodes = {}  # node_id -> node_data
    splice_node_ids = set()

    def collect_nodes(node):
        """Recursively collect endpoint nodes and splice nodes from trace tree."""
        if not node:
            return

        # Collect cable endpoint nodes
        endpoints = node.get("cable_endpoints", {})
        for key in ["start_node", "end_node"]:
            n = endpoints.get(key)
            if n and n.get("id"):
                all_endpoint_nodes[n["id"]] = n

        # Collect splice nodes (nodes where splices occur - these are in the middle)
        splice_node = node.get("node")
        if splice_node and splice_node.get("id"):
            splice_node_ids.add(splice_node["id"])

        for child in node.get("children", []):
            collect_nodes(child)

    collect_nodes(tree)

    # Terminal nodes are endpoint nodes that are NOT splice nodes
    terminal_node_ids = set(all_endpoint_nodes.keys()) - splice_node_ids
    terminal_nodes = [all_endpoint_nodes[nid] for nid in terminal_node_ids]

    # Sort by name for consistent ordering
    terminal_nodes.sort(key=lambda n: n.get("name") or "")

    start_node = None
    end_node = None

    if len(terminal_nodes) >= 2:
        start_node = {
            "id": terminal_nodes[0].get("id"),
            "name": terminal_nodes[0].get("name"),
            "address": terminal_nodes[0].get("address"),
        }
        end_node = {
            "id": terminal_nodes[-1].get("id"),
            "name": terminal_nodes[-1].get("name"),
            "address": terminal_nodes[-1].get("address"),
        }
    elif len(terminal_nodes) == 1:
        # Single terminal (dead end or loop)
        start_node = {
            "id": terminal_nodes[0].get("id"),
            "name": terminal_nodes[0].get("name"),
            "address": terminal_nodes[0].get("address"),
        }
        end_node = start_node
    elif tree and tree.get("cable_endpoints"):
        # Fallback to root cable endpoints
        endpoints = tree["cable_endpoints"]
        if endpoints.get("start_node"):
            sn = endpoints["start_node"]
            start_node = {"id": sn.get("id"), "name": sn.get("name"), "address": sn.get("address")}
        if endpoints.get("end_node"):
            en = endpoints["end_node"]
            end_node = {"id": en.get("id"), "name": en.get("name"), "address": en.get("address")}

    return {
        "fiber_id": str(fiber_id),
        "start_node": start_node,
        "end_node": end_node,
        "statistics": {
            "total_fibers": stats.get("total_fibers", 0),
            "total_splices": stats.get("total_splices", 0),
            "total_nodes": stats.get("total_nodes", 0),
            "total_addresses": stats.get("total_addresses", 0),
            "total_residential_units": stats.get("total_residential_units", 0),
        },
    }
